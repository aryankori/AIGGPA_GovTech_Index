"""
Creates the AIGGPA Fieldwork Data Vault:
- Full folder structure
- Master tracking workbook (Excel)
- Observation checklist template
- Consent form template
- README with instructions
"""
import os, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

BASE = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault"

# ── 1. Create folder structure ──
folders = [
    "01_Schedules_Raw/Revenue/HO",
    "01_Schedules_Raw/Revenue/DO",
    "01_Schedules_Raw/Rural_Development/HO",
    "01_Schedules_Raw/Rural_Development/DO",
    "01_Schedules_Raw/Forest/HO",
    "01_Schedules_Raw/Forest/DO",
    "01_Schedules_Raw/Health/HO",
    "01_Schedules_Raw/Health/DO",
    "02_Audio_Recordings/Interviews",
    "02_Audio_Recordings/FGDs",
    "03_Transcripts/Interviews",
    "03_Transcripts/FGDs",
    "04_Observation_Notes",
    "05_Photos_Evidence",
    "05_Photos_Evidence/Office_Infrastructure",
    "05_Photos_Evidence/Digital_Setup",
    "05_Photos_Evidence/Training_Material",
    "06_Consent_Forms/Scanned",
    "07_Secondary_Data/IT_Inventories",
    "07_Secondary_Data/Training_Records",
    "07_Secondary_Data/Circulars_Orders",
    "07_Secondary_Data/Helpdesk_Logs",
    "08_Data_Entry/Raw_Excel",
    "08_Data_Entry/Cleaned_Data",
    "08_Data_Entry/SPSS_Files",
    "09_Analysis_Output/Quantitative",
    "09_Analysis_Output/Qualitative_Coding",
    "09_Analysis_Output/Visualisations",
    "10_Reports_Drafts",
    "11_Presentations",
    "12_Backups",
]

for f in folders:
    os.makedirs(os.path.join(BASE, f), exist_ok=True)
print(f"Created {len(folders)} folders")

# ── Styles ──
navy = "1B2A4A"
gold = "C49A2A"
white_f = Font(color="FFFFFF", bold=True, size=11)
navy_fill = PatternFill("solid", fgColor=navy)
gold_fill = PatternFill("solid", fgColor=gold)
light_fill = PatternFill("solid", fgColor="F0F4F8")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
wrap = Alignment(wrap_text=True, vertical="top")
thin = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def style_header(ws, row, cols, fill=navy_fill):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = white_f
        cell.fill = fill
        cell.alignment = header_align
        cell.border = thin

def style_range(ws, start_row, end_row, cols):
    for r in range(start_row, end_row+1):
        for c in range(1, cols+1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin
            cell.alignment = wrap
            if r % 2 == 0:
                cell.fill = light_fill

# ── 2. Master Tracking Workbook ──
wb = openpyxl.Workbook()

# --- Sheet 1: Respondent Log ---
ws1 = wb.active
ws1.title = "Respondent_Log"
ws1.sheet_properties.tabColor = navy

h1 = ["S.No", "Respondent ID", "Date", "Department", "Office Type",
      "Cadre", "Age Group", "Gender", "Years of Service",
      "Schedule Status", "Interview?", "FGD?", "Consent?",
      "Audio File", "Photo Evidence", "Notes"]
for i, h in enumerate(h1, 1):
    ws1.cell(row=1, column=i, value=h)
style_header(ws1, 1, len(h1))

# Pre-fill 320 rows with respondent IDs
for r in range(2, 322):
    ws1.cell(row=r, column=1, value=r-1)
    ws1.cell(row=r, column=2, value=f"R{r-1:03d}")
    ws1.cell(row=r, column=10, value="Pending")
    ws1.cell(row=r, column=11, value="No")
    ws1.cell(row=r, column=12, value="No")
    ws1.cell(row=r, column=13, value="No")
style_range(ws1, 2, 321, len(h1))

# Add data validation for dropdowns
from openpyxl.worksheet.datavalidation import DataValidation
dv_dept = DataValidation(type="list", formula1='"Revenue,Rural Development,Forest,Health"')
dv_office = DataValidation(type="list", formula1='"HO,DO"')
dv_cadre = DataValidation(type="list", formula1='"Class I,Class II,Class III,Class IV"')
dv_age = DataValidation(type="list", formula1='"30-45,46-60"')
dv_gender = DataValidation(type="list", formula1='"Male,Female,Other"')
dv_status = DataValidation(type="list", formula1='"Pending,Complete,Partial,Refused"')
dv_yn = DataValidation(type="list", formula1='"Yes,No"')

ws1.add_data_validation(dv_dept)
ws1.add_data_validation(dv_office)
ws1.add_data_validation(dv_cadre)
ws1.add_data_validation(dv_age)
ws1.add_data_validation(dv_gender)
ws1.add_data_validation(dv_status)
ws1.add_data_validation(dv_yn)

dv_dept.add(f"D2:D321")
dv_office.add(f"E2:E321")
dv_cadre.add(f"F2:F321")
dv_age.add(f"G2:G321")
dv_gender.add(f"H2:H321")
dv_status.add(f"J2:J321")
dv_yn.add(f"K2:K321")
dv_yn.add(f"L2:L321")
dv_yn.add(f"M2:M321")

widths1 = [6, 14, 12, 18, 12, 12, 10, 10, 14, 14, 10, 8, 10, 25, 16, 30]
for i, w in enumerate(widths1, 1):
    ws1.column_dimensions[get_column_letter(i)].width = w

ws1.auto_filter.ref = f"A1:P321"
ws1.freeze_panes = "A2"

# --- Sheet 2: Daily Progress ---
ws2 = wb.create_sheet("Daily_Progress")
ws2.sheet_properties.tabColor = gold

h2 = ["Date", "Department Visited", "Office (HO/DO)",
      "Schedules Completed", "Interviews Done", "FGDs Done",
      "Observations Done", "Photos Taken", "Issues/Blockers", "Notes"]
for i, h in enumerate(h2, 1):
    ws2.cell(row=1, column=i, value=h)
style_header(ws2, 1, len(h2), gold_fill)

for r in range(2, 62):
    style_range(ws2, r, r, len(h2))
widths2 = [12, 20, 14, 18, 14, 12, 16, 14, 30, 30]
for i, w in enumerate(widths2, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

# --- Sheet 3: Sampling Dashboard ---
ws3 = wb.create_sheet("Sampling_Dashboard")
ws3.sheet_properties.tabColor = "2E7D32"

ws3.cell(row=1, column=1, value="SAMPLING PROGRESS TRACKER")
ws3.cell(row=1, column=1).font = Font(bold=True, size=14, color=navy)
ws3.merge_cells("A1:F1")

h3 = ["Department", "Office", "Class I", "Class II", "Class III", "Class IV", "Total"]
for i, h in enumerate(h3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, len(h3))

depts = ["Revenue", "Rural Development", "Forest", "Health"]
row = 4
for dept in depts:
    for office in ["HO", "DO"]:
        ws3.cell(row=row, column=1, value=dept)
        ws3.cell(row=row, column=2, value=office)
        for c in range(3, 7):
            ws3.cell(row=row, column=c, value=0)
        ws3.cell(row=row, column=7).value = f"=SUM(C{row}:F{row})"
        row += 1

# Target row
ws3.cell(row=row, column=1, value="TARGET")
ws3.cell(row=row, column=1).font = Font(bold=True, color="FF0000")
for c in range(3, 7):
    ws3.cell(row=row, column=c, value=80)
ws3.cell(row=row, column=7, value=320)
row += 1

# Totals
ws3.cell(row=row, column=1, value="ACTUAL TOTAL")
ws3.cell(row=row, column=1).font = Font(bold=True, color=navy)
for c in range(3, 8):
    col_letter = get_column_letter(c)
    ws3.cell(row=row, column=c).value = f"=SUM({col_letter}4:{col_letter}{row-2})"

style_range(ws3, 4, row, len(h3))
widths3 = [20, 10, 10, 10, 10, 10, 10]
for i, w in enumerate(widths3, 1):
    ws3.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 4: Interview Log ---
ws4 = wb.create_sheet("Interview_Log")
ws4.sheet_properties.tabColor = "E65100"

h4 = ["Interview #", "Respondent ID", "Date", "Department", "Cadre",
      "Duration (min)", "Audio File Name", "Consent Recorded?",
      "Key Themes (brief)", "Transcription Status"]
for i, h in enumerate(h4, 1):
    ws4.cell(row=1, column=i, value=h)
style_header(ws4, 1, len(h4))

for r in range(2, 22):
    ws4.cell(row=r, column=1, value=f"INT-{r-1:02d}")
    ws4.cell(row=r, column=10, value="Pending")
style_range(ws4, 2, 21, len(h4))
widths4 = [12, 14, 12, 18, 12, 14, 25, 16, 35, 18]
for i, w in enumerate(widths4, 1):
    ws4.column_dimensions[get_column_letter(i)].width = w
ws4.freeze_panes = "A2"

# --- Sheet 5: FGD Log ---
ws5 = wb.create_sheet("FGD_Log")
ws5.sheet_properties.tabColor = "6A1B9A"

h5 = ["FGD #", "Date", "Department", "Venue", "No. of Participants",
      "Duration (min)", "Audio File Name", "Group Consent?",
      "Key Discussion Points", "Transcription Status"]
for i, h in enumerate(h5, 1):
    ws5.cell(row=1, column=i, value=h)
style_header(ws5, 1, len(h5))

for r in range(2, 6):
    ws5.cell(row=r, column=1, value=f"FGD-{r-1:02d}")
    ws5.cell(row=r, column=10, value="Pending")
style_range(ws5, 2, 5, len(h5))
widths5 = [10, 12, 18, 20, 18, 14, 25, 16, 35, 18]
for i, w in enumerate(widths5, 1):
    ws5.column_dimensions[get_column_letter(i)].width = w

# --- Sheet 6: Observation Log ---
ws6 = wb.create_sheet("Observation_Log")
ws6.sheet_properties.tabColor = "00695C"

h6 = ["Visit #", "Date", "Department", "Office (HO/DO)", "Location",
      "Computers Visible", "Computers In Use", "Internet Working?",
      "Internet Speed (if tested)", "Paper vs Digital Ratio",
      "IT Helpdesk Present?", "Personal Phones for Work?",
      "UPS/Power Backup?", "Photo File Names", "Field Notes"]
for i, h in enumerate(h6, 1):
    ws6.cell(row=1, column=i, value=h)
style_header(ws6, 1, len(h6))

for r in range(2, 22):
    ws6.cell(row=r, column=1, value=f"OBS-{r-1:02d}")
style_range(ws6, 2, 21, len(h6))
widths6 = [10, 12, 18, 14, 20, 16, 16, 14, 18, 18, 16, 18, 14, 25, 30]
for i, w in enumerate(widths6, 1):
    ws6.column_dimensions[get_column_letter(i)].width = w
ws6.freeze_panes = "A2"

# --- Sheet 7: File Naming Convention ---
ws7 = wb.create_sheet("File_Naming_Guide")
ws7.sheet_properties.tabColor = "37474F"

guide = [
    ["File Type", "Naming Convention", "Example", "Store In"],
    ["Scanned Schedule", "SCH_[RespID]_[Date].pdf", "SCH_R045_2026-05-15.pdf", "01_Schedules_Raw/[Dept]/[Office]"],
    ["Interview Audio", "INT_[RespID]_[Date].m4a", "INT_R045_2026-05-15.m4a", "02_Audio_Recordings/Interviews"],
    ["FGD Audio", "FGD_[Dept]_[Date].m4a", "FGD_Revenue_2026-05-20.m4a", "02_Audio_Recordings/FGDs"],
    ["Interview Transcript", "TRANS_INT_[RespID].docx", "TRANS_INT_R045.docx", "03_Transcripts/Interviews"],
    ["FGD Transcript", "TRANS_FGD_[Dept].docx", "TRANS_FGD_Revenue.docx", "03_Transcripts/FGDs"],
    ["Observation Notes", "OBS_[Dept]_[Office]_[Date].pdf", "OBS_Revenue_HO_2026-05-15.pdf", "04_Observation_Notes"],
    ["Office Photo", "PHOTO_[Dept]_[Office]_[Desc]_[Date].jpg", "PHOTO_Revenue_HO_Server_2026-05-15.jpg", "05_Photos_Evidence"],
    ["Consent Form", "CONSENT_[RespID].pdf", "CONSENT_R045.pdf", "06_Consent_Forms/Scanned"],
    ["Dept Circular", "CIRC_[Dept]_[Desc]_[Year].pdf", "CIRC_Revenue_eOffice_Mandate_2024.pdf", "07_Secondary_Data/Circulars_Orders"],
    ["Data Entry File", "DATA_MASTER_[Date].xlsx", "DATA_MASTER_2026-06-01.xlsx", "08_Data_Entry/Raw_Excel"],
    ["SPSS File", "AIGGPA_DATA_CLEAN.sav", "AIGGPA_DATA_CLEAN.sav", "08_Data_Entry/SPSS_Files"],
]

for r, row_data in enumerate(guide, 1):
    for c, val in enumerate(row_data, 1):
        ws7.cell(row=r, column=c, value=val)

style_header(ws7, 1, 4)
style_range(ws7, 2, len(guide), 4)
for i, w in enumerate([20, 35, 35, 40], 1):
    ws7.column_dimensions[get_column_letter(i)].width = w

# Save workbook
tracker_path = os.path.join(BASE, "AIGGPA_Master_Tracker.xlsx")
wb.save(tracker_path)
print(f"Created: AIGGPA_Master_Tracker.xlsx (7 sheets)")

# ── 3. Observation Checklist Template ──
obs_text = """AIGGPA OFFICE OBSERVATION CHECKLIST
====================================
Visit #: ______    Date: ___/___/2026    Time: ____:____ to ____:____

Department: [ ] Revenue  [ ] Rural Dev  [ ] Forest  [ ] Health
Office Type: [ ] Head Office  [ ] District Office
Location/Address: _________________________________________________
Officer Met: ______________________  Designation: __________________

SECTION A: HARDWARE & DEVICES
1. Desktop computers visible:        Count: ____
2. Laptops visible:                  Count: ____
3. Printers available:               Count: ____
4. Scanners available:               Count: ____
5. Devices currently in use:         [ ] All  [ ] Most  [ ] Some  [ ] None
6. Shared devices observed:          [ ] Yes  [ ] No
7. Device condition:                 [ ] Good  [ ] Fair  [ ] Poor

SECTION B: CONNECTIVITY
8. Internet available:               [ ] Yes  [ ] No
9. Speed test (if possible):         Download: ____Mbps  Upload: ____Mbps
10. WiFi or LAN:                     [ ] WiFi  [ ] LAN  [ ] Both  [ ] Neither
11. Power backup (UPS/Inverter):     [ ] Yes  [ ] No
12. Power outage during visit:       [ ] Yes  [ ] No

SECTION C: DIGITAL vs PAPER
13. Paper files on desks:            [ ] Heavy  [ ] Moderate  [ ] Minimal
14. Registers/ledgers in active use: [ ] Yes  [ ] No
15. Employees typing on screens:     [ ] Most  [ ] Some  [ ] Few  [ ] None
16. Parallel paper+digital observed: [ ] Yes  [ ] No

SECTION D: SUPPORT & ENVIRONMENT
17. IT helpdesk/support visible:     [ ] Yes  [ ] No
18. IT support contact displayed:    [ ] Yes  [ ] No
19. Employees using personal phones for work: [ ] Yes  [ ] No
20. Government portal open on any screen:     [ ] Yes  [ ] No
    If yes, which portal: _________________________________________

SECTION E: GENERAL NOTES
21. Overall digital readiness impression: [ ] High  [ ] Medium  [ ] Low
22. Notable observations:
___________________________________________________________________
___________________________________________________________________
___________________________________________________________________
23. Photos taken: [ ] Yes (filenames: _____________________________) [ ] No

Observer Signature: ________________    Date: ___/___/2026
"""

obs_path = os.path.join(BASE, "04_Observation_Notes", "TEMPLATE_Observation_Checklist.txt")
with open(obs_path, "w", encoding="utf-8") as f:
    f.write(obs_text)
print("Created: Observation checklist template")

# ── 4. Consent Form Template ──
consent = """INFORMED CONSENT FORM
=====================
Study: Assessment of the Use of Digital Tools and Technologies by
       Government Employees for Enhancing Workplace Efficiency
       and Effectiveness

Institution: Atal Bihari Vajpayee Institute of Good Governance
             and Policy Analysis (AIGGPA), Bhopal

Researcher: Aryan Kori (Research Intern)

---------------------------------------------------------------

Dear Participant,

You are invited to participate in a research study being conducted
under AIGGPA. Your participation is entirely voluntary.

PURPOSE: To understand how government employees use digital tools
in their daily work and identify areas for improvement.

WHAT IS INVOLVED:
- A structured schedule (approximately 15-20 minutes)
- You may also be invited for a brief interview or group discussion

CONFIDENTIALITY:
- Your responses will be kept strictly confidential
- No individual will be identified in the final report
- Data will be used for research purposes only

AUDIO RECORDING (if applicable):
[ ] I consent to audio recording of my interview/FGD
[ ] I do NOT consent to audio recording

YOUR RIGHTS:
- You may refuse to answer any question
- You may withdraw at any time without consequence
- You may request your data to be removed from the study

---------------------------------------------------------------

PARTICIPANT DECLARATION:

I have read and understood the above information. I voluntarily
agree to participate in this study.

Name: ___________________________________________

Designation: ____________________________________

Department: _____________________________________

Signature: ________________    Date: ___/___/2026

---------------------------------------------------------------

RESEARCHER DECLARATION:

I have explained the study to the participant and answered their
questions. I believe the participant has given informed consent.

Researcher Signature: ________________   Date: ___/___/2026

Respondent ID (office use): ___________
"""

consent_path = os.path.join(BASE, "06_Consent_Forms", "TEMPLATE_Consent_Form.txt")
with open(consent_path, "w", encoding="utf-8") as f:
    f.write(consent)
print("Created: Consent form template")

# ── 5. README ──
readme = """# AIGGPA FIELDWORK DATA VAULT
## Directory Structure & Usage Guide

Created: """ + datetime.now().strftime("%Y-%m-%d") + """

## Folder Structure

```
AIGGPA_Fieldwork_Vault/
|
|-- AIGGPA_Master_Tracker.xlsx    <-- YOUR COMMAND CENTER (open this first)
|
|-- 01_Schedules_Raw/             <-- Scanned/photographed filled schedules
|   |-- Revenue/HO/              <-- Revenue Head Office schedules
|   |-- Revenue/DO/              <-- Revenue District Office schedules
|   |-- Rural_Development/HO/
|   |-- Rural_Development/DO/
|   |-- Forest/HO/
|   |-- Forest/DO/
|   |-- Health/HO/
|   |-- Health/DO/
|
|-- 02_Audio_Recordings/          <-- All audio evidence
|   |-- Interviews/              <-- Individual interview recordings
|   |-- FGDs/                    <-- Focus group recordings
|
|-- 03_Transcripts/               <-- Typed transcriptions
|   |-- Interviews/
|   |-- FGDs/
|
|-- 04_Observation_Notes/         <-- Filled observation checklists
|   |-- TEMPLATE_Observation_Checklist.txt
|
|-- 05_Photos_Evidence/           <-- Photographic proof
|   |-- Office_Infrastructure/   <-- Office setup, hardware
|   |-- Digital_Setup/           <-- Screens, portals in use
|   |-- Training_Material/       <-- Training schedules, certificates
|
|-- 06_Consent_Forms/             <-- Signed consent records
|   |-- Scanned/                 <-- Scanned PDFs of signed forms
|   |-- TEMPLATE_Consent_Form.txt
|
|-- 07_Secondary_Data/            <-- Documents collected from departments
|   |-- IT_Inventories/          <-- Hardware/software lists
|   |-- Training_Records/        <-- Training attendance, schedules
|   |-- Circulars_Orders/        <-- Official orders on digital tools
|   |-- Helpdesk_Logs/           <-- IT complaint records
|
|-- 08_Data_Entry/                <-- Your analysis-ready data
|   |-- Raw_Excel/               <-- Daily data entry spreadsheets
|   |-- Cleaned_Data/            <-- Final cleaned dataset
|   |-- SPSS_Files/              <-- .sav files for SPSS analysis
|
|-- 09_Analysis_Output/           <-- Results
|   |-- Quantitative/           <-- SPSS output, tables, stats
|   |-- Qualitative_Coding/     <-- Thematic analysis codebooks
|   |-- Visualisations/         <-- Charts, graphs, heatmaps
|
|-- 10_Reports_Drafts/            <-- Report drafts
|-- 11_Presentations/             <-- PPTs for stakeholder briefings
|-- 12_Backups/                   <-- Weekly backup copies
```

## Master Tracker Sheets

1. **Respondent_Log** -- One row per respondent (R001-R320). Track status.
2. **Daily_Progress** -- Fill every evening. What you did today.
3. **Sampling_Dashboard** -- Live count of how many per cell you've completed.
4. **Interview_Log** -- Track all 12-16 interviews.
5. **FGD_Log** -- Track all 4 FGDs.
6. **Observation_Log** -- Track every office visit observation.
7. **File_Naming_Guide** -- How to name every file consistently.

## Daily Routine

1. Collect data in the field
2. Come back, scan/photograph filled schedules
3. Enter data into the master data entry sheet
4. Update the Respondent_Log (mark Complete)
5. Update Daily_Progress
6. Update Sampling_Dashboard counts
7. Back up the tracker to 12_Backups/ every Friday

## File Naming Rules

- Always use the Respondent ID (R001, R002...) in filenames
- Always include the date as YYYY-MM-DD
- No spaces in filenames -- use underscores
- See the File_Naming_Guide sheet for exact patterns
"""

readme_path = os.path.join(BASE, "README.md")
with open(readme_path, "w", encoding="utf-8") as f:
    f.write(readme)
print("Created: README.md")

print(f"\nDone! Vault created at:\n{BASE}")
