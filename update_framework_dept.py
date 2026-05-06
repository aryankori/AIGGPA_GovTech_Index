import openpyxl

def update_framework():
    filepath = r"C:\Users\aryan\Downloads\Framework for Interns.xlsx"
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Update Objective 1 (Row 3)
    current_q1 = ws.cell(row=3, column=8).value
    if current_q1 and "[Dept-Specific Supplement]" not in current_q1:
        ws.cell(row=3, column=8).value = current_q1 + "\n\n[Dept-Specific Supplement:]\n- Revenue: Awareness of RCMS, SAMPADA 2.0\n- Rural Dev: Awareness of NMMS, e-Gram Swaraj\n- Forest: Awareness of e-Green Watch, AI Alerts\n- Health: Awareness of ANMOL, eVIN, IHIP"
        
    # Update Objective 3 (Row 5)
    current_q3 = ws.cell(row=5, column=8).value
    if current_q3 and "[Dept-Specific Supplement]" not in current_q3:
        ws.cell(row=5, column=8).value = current_q3 + "\n\n[Dept-Specific Supplement:]\n- Revenue: Difficulty managing legacy paper vs digital records\n- Rural Dev: Burden of multi-portal data entry\n- Forest: Challenges using GIS & poor network in remote areas\n- Health: Dual documentation burden (paper + HMIS)"

    # Update Objective 4 (Row 6)
    current_q4 = ws.cell(row=6, column=8).value
    if current_q4 and "[Dept-Specific Supplement]" not in current_q4:
        ws.cell(row=6, column=8).value = current_q4 + "\n\n[Dept-Specific Supplement:]\n- Impact of specific flagship tools (e.g., Does AI monitoring actually stop forest offenses? Does ANMOL actually improve patient follow-up?)"

    wb.save(filepath)
    print("Updated Framework for Interns.xlsx with department-specific questions.")

if __name__ == "__main__":
    update_framework()
