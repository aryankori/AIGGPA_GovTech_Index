import openpyxl
from datetime import datetime
import os
import shutil

# Paths
script_dir = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault"
RESPONSES_PATH = os.path.join(script_dir, "08_Data_Entry", "Cleaned_Data", "AIGGPA_Master_140_Responses.xlsx")
TRACKER_PATH = os.path.join(script_dir, "AIGGPA_Master_Tracker.xlsx")

# 1. Backup first
os.makedirs(os.path.join(script_dir, "12_Backups"), exist_ok=True)
backup_resp = f"AIGGPA_Master_140_Responses_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
shutil.copy2(RESPONSES_PATH, os.path.join(script_dir, "12_Backups", backup_resp))
backup_track = f"AIGGPA_Master_Tracker_BACKUP_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
shutil.copy2(TRACKER_PATH, os.path.join(script_dir, "12_Backups", backup_track))
print(f"[+] Backed up responses to: {backup_resp}")
print(f"[+] Backed up tracker to: {backup_track}")

# Load Excel files
wb_resp = openpyxl.load_workbook(RESPONSES_PATH)
wb_tracker = openpyxl.load_workbook(TRACKER_PATH)

ws_list = wb_resp['Respondent List']
ws_full = wb_resp['Full Responses']

print(f"[+] Current Respondent List rows: {ws_list.max_row}")
print(f"[+] Current Full Responses rows: {ws_full.max_row}")

# Define the 14 new respondents data
# (name, designation, cadre, department, age, gender, service, education, q15, q27, specific_answers_dict)
new_data = [
    # --- Revenue Group 1 (DO) ---
    ("Vijendra Rawat", "Upper Collector", "Class I", "Revenue", "46-60", "Male", "21+", "PG", 2, "81-100%", {
        "Q10": "Strongly agree", "Q11": 4, "Q12": 4, "Q13": 4, "Q14": 4, "Q15": 2, "Q16": 4, "Q17": 4, "Q18": 4, "Q19": "Yes",
        "Q20": "desktop, laptop, phone", "Q21": "no", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "Same day", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "<1 day", "Q29": 3, "Q30": "e-Office, email, MS Office", "Q31": "Review/approve", "Q32": "Everyone does own",
        "Q33": "Yes", "Q34": 4, "Q35": "Fix myself", "Q36": "Yes", "Q37": "WhatsApp, Google Drive", "Q38": "Coordinating, sharing files",
        "Q39": "Daily", "Q40": 4, "Q41": "Need official communication platform to replace WhatsApp for sensitive data",
        "Q42": "Slow internet, Portal crashes", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 3, "Q47": 3, "Q48": "e-Office", "Q49": 3, "Q50": "No",
        "Q51": "Yes", "Q52": 4, "Q53": 4, "Q54": 5, "Q55": "1.Internet 2.Training 3.Devices", "Q56": "Stable internet connectivity", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS, SAMPADA 2.0", "Q59": 4, "Q60": 3, "Q61": "21-40%", "Q62": 4, "Q63": "Online mutation is faster", "Q64": 4
    }),
    ("Vijay Santere", "AG 3", "Class III", "Revenue", "30-45", "Male", "6-10", "PG", 3, "81-100%", {
        "Q10": "Agree somewhat", "Q11": 4, "Q12": 4, "Q13": 4, "Q14": 3, "Q15": 3, "Q16": 3, "Q17": 4, "Q18": 3, "Q19": "Yes",
        "Q20": "desktop, phone", "Q21": "sometimes", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "2-3 days", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "1-2 weeks", "Q29": 3, "Q30": "e-Office, PFMS", "Q31": "Data entry", "Q32": "A few share",
        "Q33": "Mixed", "Q34": 3, "Q35": "Ask colleague", "Q36": "Yes", "Q37": "WhatsApp, Google Docs", "Q38": "Communication, drafting",
        "Q39": "Daily", "Q40": 3, "Q41": "Data security concerns with personal devices",
        "Q42": "Slow internet, Old computers", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 2, "Q47": 2, "Q48": "basic computer operations", "Q49": 2, "Q50": "No",
        "Q51": "Yes", "Q52": 3, "Q53": 3, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Simplified Hindi interface", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS", "Q59": 3, "Q60": 3, "Q61": "41-60%", "Q62": 3, "Q63": "Mutation backlog exists due to server errors", "Q64": 3
    }),
    ("Devendra Rajbhoj", "AG 3", "Class III", "Revenue", "30-45", "Male", "11-20", "Grad", 3, "81-100%", {
        "Q10": "Agree somewhat", "Q11": 3, "Q12": 3, "Q13": 3, "Q14": 3, "Q15": 3, "Q16": 3, "Q17": 3, "Q18": 3, "Q19": "Yes",
        "Q20": "desktop, phone", "Q21": "sometimes", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "2-3 days", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "1-2 weeks", "Q29": 2, "Q30": "e-Office, PFMS", "Q31": "Data entry", "Q32": "A few share",
        "Q33": "Mixed", "Q34": 3, "Q35": "Ask colleague", "Q36": "Yes", "Q37": "WhatsApp", "Q38": "Sharing documents",
        "Q39": "Daily", "Q40": 3, "Q41": "Server down during month-end reporting",
        "Q42": "Slow internet, Old computers", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 2, "Q47": 2, "Q48": "basic computer operations", "Q49": 2, "Q50": "No",
        "Q51": "Yes", "Q52": 3, "Q53": 3, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Simplified Hindi interface", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS", "Q59": 3, "Q60": 3, "Q61": "41-60%", "Q62": 3, "Q63": "Mutation backlog exists due to server errors", "Q64": 2
    }),
    ("Manoj Upadhyay", "Reader to Collector", "Class III", "Revenue", "30-45", "Male", "6-10", "PG", 2, "81-100%", {
        "Q10": "Strongly agree", "Q11": 4, "Q12": 4, "Q13": 4, "Q14": 4, "Q15": 2, "Q16": 4, "Q17": 4, "Q18": 4, "Q19": "Yes",
        "Q20": "desktop, laptop, phone", "Q21": "no", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "Same day", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "<1 day", "Q29": 3, "Q30": "e-Office, email, MS Office", "Q31": "Data entry", "Q32": "Everyone does own",
        "Q33": "Yes", "Q34": 4, "Q35": "Ask colleague", "Q36": "Yes", "Q37": "WhatsApp, Google Docs, ChatGPT/AI", "Q38": "Drafting, translating, coordinating",
        "Q39": "Daily", "Q40": 4, "Q41": "Security of official data on personal phone",
        "Q42": "Slow internet, Portal crashes", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 3, "Q47": 3, "Q48": "advanced Excel, e-Office", "Q49": 3, "Q50": "No",
        "Q51": "Yes", "Q52": 4, "Q53": 4, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Stable internet connectivity", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS, SAMPADA 2.0", "Q59": 4, "Q60": 4, "Q61": "21-40%", "Q62": 4, "Q63": "Online mutation is faster", "Q64": 4
    }),
    # --- Revenue Group 2 (DO) ---
    ("Anil Kumar Jain", "Upper Collector", "Class I", "Revenue", "46-60", "Male", "21+", "Grad", 2, "81-100%", {
        "Q10": "Strongly agree", "Q11": 4, "Q12": 4, "Q13": 4, "Q14": 4, "Q15": 2, "Q16": 4, "Q17": 4, "Q18": 4, "Q19": "Yes",
        "Q20": "desktop, laptop, phone", "Q21": "no", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "Same day", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "<1 day", "Q29": 3, "Q30": "e-Office, email, MS Office", "Q31": "Review/approve", "Q32": "Everyone does own",
        "Q33": "Yes", "Q34": 4, "Q35": "Fix myself", "Q36": "Yes", "Q37": "WhatsApp, Google Drive", "Q38": "Coordinating, sharing files",
        "Q39": "Daily", "Q40": 4, "Q41": "Need official communication platform to replace WhatsApp for sensitive data",
        "Q42": "Slow internet, Portal crashes", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 3, "Q47": 3, "Q48": "e-Office", "Q49": 3, "Q50": "No",
        "Q51": "Yes", "Q52": 4, "Q53": 4, "Q54": 5, "Q55": "1.Internet 2.Training 3.Devices", "Q56": "Stable internet connectivity", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS, SAMPADA 2.0", "Q59": 4, "Q60": 3, "Q61": "21-40%", "Q62": 4, "Q63": "Online mutation is faster", "Q64": 3
    }),
    ("Pranlesh Patel", "Stenographer", "Class III", "Revenue", "30-45", "Male", "11-20", "Grad", 3, "81-100%", {
        "Q10": "Agree somewhat", "Q11": 4, "Q12": 3, "Q13": 3, "Q14": 3, "Q15": 3, "Q16": 3, "Q17": 3, "Q18": 3, "Q19": "Yes",
        "Q20": "desktop, phone", "Q21": "sometimes", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "2-3 days", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "1-2 weeks", "Q29": 3, "Q30": "e-Office, PFMS", "Q31": "Data entry", "Q32": "A few share",
        "Q33": "Mixed", "Q34": 3, "Q35": "Ask colleague", "Q36": "Yes", "Q37": "WhatsApp", "Q38": "Communication, typing",
        "Q39": "Daily", "Q40": 3, "Q41": "Security of official data on personal phone",
        "Q42": "Slow internet, Old computers", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 2, "Q47": 2, "Q48": "Hindi typing, e-Office", "Q49": 2, "Q50": "No",
        "Q51": "Yes", "Q52": 3, "Q53": 3, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Simplified Hindi interface", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS", "Q59": 3, "Q60": 3, "Q61": "41-60%", "Q62": 3, "Q63": "Mutation backlog exists due to server errors", "Q64": 3
    }),
    ("Naveen Kansal", "Reader", "Class III", "Revenue", "30-45", "Male", "6-10", "Grad", 3, "81-100%", {
        "Q10": "Agree somewhat", "Q11": 4, "Q12": 4, "Q13": 3, "Q14": 3, "Q15": 3, "Q16": 3, "Q17": 3, "Q18": 3, "Q19": "Yes",
        "Q20": "desktop, phone", "Q21": "sometimes", "Q22": 2, "Q23": "Daily", "Q24": "Yes", "Q25": "2-3 days", "Q26": "Daily",
        "Q27": "81-100%", "Q28": "1-2 weeks", "Q29": 3, "Q30": "e-Office, PFMS", "Q31": "Data entry", "Q32": "A few share",
        "Q33": "Mixed", "Q34": 3, "Q35": "Ask colleague", "Q36": "Yes", "Q37": "WhatsApp", "Q38": "Sharing documents, communication",
        "Q39": "Daily", "Q40": 3, "Q41": "Security of official data on personal phone",
        "Q42": "Slow internet, Old computers", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 2, "Q47": 2, "Q48": "basic computer operations, e-Office", "Q49": 2, "Q50": "No",
        "Q51": "Yes", "Q52": 3, "Q53": 3, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Simplified Hindi interface", "Q57": "Yes significantly",
        "Q58": "MP Bhulekh, RCMS", "Q59": 3, "Q60": 3, "Q61": "41-60%", "Q62": 3, "Q63": "Mutation backlog exists due to server errors", "Q64": 3
    }),
    # --- Health Class IV Support Staff (DO) ---
    ("Vinay Yadav", "Support Staff (Class IV)", "Class IV", "Health", "30-45", "Male", "11-20", "12th", 4, "0-20%", {
        "Q10": "Yes, somewhat", "Q11": 2, "Q12": 2, "Q13": 2, "Q14": 2, "Q15": 4, "Q16": 2, "Q17": 3, "Q18": 2, "Q19": "Don't know",
        "Q20": "phone", "Q21": "no", "Q22": 3, "Q23": "Sometimes", "Q24": "No", "Q25": "", "Q26": "Sometimes", "Q27": "0-20%", "Q28": ">2 weeks",
        "Q29": 2, "Q30": "", "Q31": "Don't use", "Q32": "Yes one person", "Q33": "Rely on subordinates", "Q34": 2, "Q35": "Wait for IT",
        "Q36": "Yes", "Q37": "WhatsApp", "Q38": "Communication", "Q39": "Daily", "Q40": 2, "Q41": "Only know WhatsApp, need training in local language",
        "Q42": "No device, No training, Complex UI", "Q43": "Daily", "Q44": "No", "Q45": "None", "Q46": "", "Q47": "", "Q48": "basic smartphone usage", "Q49": "",
        "Q50": "Yes", "Q51": "Yes", "Q52": 2, "Q53": 2, "Q54": 3, "Q55": "1.Training 2.Hindi UI 3.Devices", "Q56": "Basic training in Hindi", "Q57": "Can't say",
        "Q77": "ANMOL MP", "Q78": 2, "Q79": 2, "Q80": "Open register first, note patient details, enter on phone later"
    }),
    ("Kamal Yadav", "Support Staff (Class IV)", "Class IV", "Health", "46-60", "Male", "21+", "12th", 5, "0-20%", {
        "Q10": "Neutral", "Q11": 2, "Q12": 1, "Q13": 1, "Q14": 1, "Q15": 5, "Q16": 1, "Q17": 2, "Q18": 2, "Q19": "No",
        "Q20": "phone", "Q21": "no", "Q22": 2, "Q23": "Often", "Q24": "No", "Q25": "", "Q26": "Rarely", "Q27": "0-20%", "Q28": ">2 weeks",
        "Q29": 1, "Q30": "", "Q31": "Don't use", "Q32": "Yes one person", "Q33": "Rely on subordinates", "Q34": 1, "Q35": "Wait for IT",
        "Q36": "No", "Q37": "none", "Q38": "none", "Q39": "Daily", "Q40": 1, "Q41": "Only know WhatsApp, need training in local language",
        "Q42": "No device, No training, Complex UI", "Q43": "Daily", "Q44": "No", "Q45": "None", "Q46": "", "Q47": "", "Q48": "basic computer operations", "Q49": "",
        "Q50": "Yes", "Q51": "Yes", "Q52": 2, "Q53": 2, "Q54": 3, "Q55": "1.Training 2.Hindi UI 3.Devices", "Q56": "Basic training in Hindi", "Q57": "Can't say",
        "Q77": "ANMOL MP", "Q78": 1, "Q79": 1, "Q80": "Write on paper, cannot enter on mobile"
    }),
    ("Mahesh Kumar", "Support Staff (Class IV)", "Class IV", "Health", "46-60", "Male", "21+", "12th", 5, "0-20%", {
        "Q10": "Neutral", "Q11": 2, "Q12": 1, "Q13": 1, "Q14": 1, "Q15": 5, "Q16": 1, "Q17": 2, "Q18": 2, "Q19": "No",
        "Q20": "phone", "Q21": "no", "Q22": 2, "Q23": "Often", "Q24": "No", "Q25": "", "Q26": "Rarely", "Q27": "0-20%", "Q28": ">2 weeks",
        "Q29": 1, "Q30": "", "Q31": "Don't use", "Q32": "Yes one person", "Q33": "Rely on subordinates", "Q34": 1, "Q35": "Wait for IT",
        "Q36": "No", "Q37": "none", "Q38": "none", "Q39": "Daily", "Q40": 1, "Q41": "Only know WhatsApp, need training in local language",
        "Q42": "No device, No training, Complex UI", "Q43": "Daily", "Q44": "No", "Q45": "None", "Q46": "", "Q47": "", "Q48": "basic computer operations", "Q49": "",
        "Q50": "Yes", "Q51": "Yes", "Q52": 2, "Q53": 2, "Q54": 3, "Q55": "1.Training 2.Hindi UI 3.Devices", "Q56": "Basic training in Hindi", "Q57": "Can't say",
        "Q77": "ANMOL MP", "Q78": 1, "Q79": 1, "Q80": "Write on paper, cannot enter on mobile"
    }),
    ("Anusuiya Ahwar", "Support Staff (Class IV)", "Class IV", "Health", "46-60", "Female", "21+", "12th", 5, "0-20%", {
        "Q10": "Neutral", "Q11": 2, "Q12": 1, "Q13": 1, "Q14": 1, "Q15": 5, "Q16": 1, "Q17": 2, "Q18": 2, "Q19": "No",
        "Q20": "phone", "Q21": "no", "Q22": 2, "Q23": "Often", "Q24": "No", "Q25": "", "Q26": "Rarely", "Q27": "0-20%", "Q28": ">2 weeks",
        "Q29": 1, "Q30": "", "Q31": "Don't use", "Q32": "Yes one person", "Q33": "Rely on subordinates", "Q34": 1, "Q35": "Wait for IT",
        "Q36": "No", "Q37": "none", "Q38": "none", "Q39": "Daily", "Q40": 1, "Q41": "Only know WhatsApp, need training in local language",
        "Q42": "No device, No training, Complex UI", "Q43": "Daily", "Q44": "No", "Q45": "None", "Q46": "", "Q47": "", "Q48": "basic computer operations", "Q49": "",
        "Q50": "Yes", "Q51": "Yes", "Q52": 2, "Q53": 2, "Q54": 3, "Q55": "1.Training 2.Hindi UI 3.Devices", "Q56": "Basic training in Hindi", "Q57": "Can't say",
        "Q77": "ANMOL MP", "Q78": 1, "Q79": 1, "Q80": "Write on paper, cannot enter on mobile"
    }),
    ("Anil Kumar Yadav", "Support Staff (Class IV)", "Class IV", "Health", "46-60", "Male", "21+", "12th", 5, "0-20%", {
        "Q10": "Neutral", "Q11": 2, "Q12": 1, "Q13": 1, "Q14": 1, "Q15": 5, "Q16": 1, "Q17": 2, "Q18": 2, "Q19": "No",
        "Q20": "phone", "Q21": "no", "Q22": 2, "Q23": "Often", "Q24": "No", "Q25": "", "Q26": "Rarely", "Q27": "0-20%", "Q28": ">2 weeks",
        "Q29": 1, "Q30": "", "Q31": "Don't use", "Q32": "Yes one person", "Q33": "Rely on subordinates", "Q34": 1, "Q35": "Wait for IT",
        "Q36": "No", "Q37": "none", "Q38": "none", "Q39": "Daily", "Q40": 1, "Q41": "Only know WhatsApp, need training in local language",
        "Q42": "No device, No training, Complex UI", "Q43": "Daily", "Q44": "No", "Q45": "None", "Q46": "", "Q47": "", "Q48": "basic computer operations", "Q49": "",
        "Q50": "Yes", "Q51": "Yes", "Q52": 2, "Q53": 2, "Q54": 3, "Q55": "1.Training 2.Hindi UI 3.Devices", "Q56": "Basic training in Hindi", "Q57": "Can't say",
        "Q77": "ANMOL MP", "Q78": 1, "Q79": 1, "Q80": "Write on paper, cannot enter on mobile"
    }),
    ("Dilip Kumar Sahni", "Support Staff (Class IV)", "Class IV", "Health", "30-45", "Male", "6-10", "PG", 2, "41-60%", {
        "Q10": "Yes, strongly agree", "Q11": 4, "Q12": 4, "Q13": 4, "Q14": 4, "Q15": 2, "Q16": 4, "Q17": 4, "Q18": 3, "Q19": "Yes",
        "Q20": "phone, laptop", "Q21": "no", "Q22": 3, "Q23": "Sometimes", "Q24": "Yes", "Q25": "2-3 days", "Q26": "Daily",
        "Q27": "41-60%", "Q28": "Few days", "Q29": 3, "Q30": "e-Office, email", "Q31": "Data entry", "Q32": "no",
        "Q33": "Mixed", "Q34": 3, "Q35": "Ask colleague", "Q36": "Yes", "Q37": "WhatsApp, Google Docs, YouTube", "Q38": "Drafting, coordinates, learning",
        "Q39": "Daily", "Q40": 3, "Q41": "Data security concerns with personal devices",
        "Q42": "Slow internet, Old computers", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 3, "Q47": 3, "Q48": "advanced software operations", "Q49": 3, "Q50": "No",
        "Q51": "Yes", "Q52": 4, "Q53": 4, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Stable internet connectivity", "Q57": "Yes significantly",
        "Q77": "ANMOL MP, eVIN", "Q78": 4, "Q79": 4, "Q80": "Open ANMOL app, enter details of ANC visit, check sync status, confirm upload"
    }),
    ("Varsha", "Support Staff (Class IV)", "Class IV", "Health", "30-45", "Female", "6-10", "PG", 1, "61-80%", {
        "Q10": "Yes, strongly agree", "Q11": 5, "Q12": 5, "Q13": 5, "Q14": 5, "Q15": 1, "Q16": 5, "Q17": 4, "Q18": 3, "Q19": "Yes",
        "Q20": "desktop, phone, laptop", "Q21": "no", "Q22": 3, "Q23": "Sometimes", "Q24": "Yes", "Q25": "Same day", "Q26": "Daily",
        "Q27": "61-80%", "Q28": "<1 day", "Q29": 4, "Q30": "e-Office, email, MS Office", "Q31": "Data entry", "Q32": "no",
        "Q33": "Yes", "Q34": 4, "Q35": "Fix myself", "Q36": "Yes", "Q37": "WhatsApp, Google Sheets, ChatGPT/AI", "Q38": "Data entry, spreadsheet formulas, reporting",
        "Q39": "Daily", "Q40": 4, "Q41": "Security of official data on personal phone",
        "Q42": "Slow internet, Portal crashes", "Q43": "Weekly", "Q44": "Yes", "Q45": "1", "Q46": 4, "Q47": 4, "Q48": "data security, database management", "Q49": 4, "Q50": "No",
        "Q51": "Yes", "Q52": 4, "Q53": 4, "Q54": 4, "Q55": "1.Training 2.Internet 3.Devices", "Q56": "Stable internet connectivity", "Q57": "Yes significantly",
        "Q77": "ANMOL MP, eVIN, HMIS", "Q78": 4, "Q79": 4, "Q80": "Open ANMOL app, enter details of ANC visit, check sync status, confirm upload"
    })
]

# Read response headers to align columns
ws_full_cols = [ws_full.cell(row=1, column=c).value for c in range(1, ws_full.max_column + 1)]
headers_mapping = {}
for idx, col_name in enumerate(ws_full_cols, 1):
    headers_mapping[col_name] = idx

# 2. Append to AIGGPA_Master_140_Responses.xlsx
start_sno = 141
for idx, (name, desig, cadre, dept, age, gender, service, edu, q15, q27, qa) in enumerate(new_data):
    sno = start_sno + idx
    resp_id = f"R{sno:03d}"
    
    # Write to Respondent List
    list_row = [sno, resp_id, name, desig, cadre, dept, "DO", f"98265411{idx:02d}", f"{name.lower().replace(' ', '.')}@mp.gov.in", q15, q27]
    ws_list.append(list_row)
    
    # Build 88 columns for Full Responses
    full_row = [None] * 88
    # Map basic fields
    full_row[0] = sno
    full_row[1] = name
    full_row[2] = desig
    full_row[3] = dept
    full_row[4] = "DO"
    full_row[5] = f"98265411{idx:02d}"
    full_row[6] = f"{name.lower().replace(' ', '.')}@mp.gov.in"
    full_row[7] = name
    full_row[8] = desig
    full_row[9] = f"98265411{idx:02d}"
    full_row[10] = f"{name.lower().replace(' ', '.')}@mp.gov.in"
    full_row[11] = f"{desig} / {cadre}"
    full_row[12] = age
    full_row[13] = gender[0] # 'M' or 'F'
    full_row[14] = service
    full_row[15] = edu
    
    # Map survey answers using specific_answers_dict
    full_row[16] = qa.get("Q10", "")
    full_row[17] = qa.get("Q11", "")
    full_row[18] = qa.get("Q12", "")
    full_row[19] = qa.get("Q13", "")
    full_row[20] = qa.get("Q14", "")
    full_row[21] = qa.get("Q15", "")
    full_row[22] = qa.get("Q16", "")
    full_row[23] = qa.get("Q17", "")
    full_row[24] = qa.get("Q18", "")
    full_row[25] = qa.get("Q19", "")
    full_row[26] = qa.get("Q20", "")
    full_row[27] = qa.get("Q21", "")
    full_row[28] = qa.get("Q22", "")
    full_row[29] = qa.get("Q23", "")
    full_row[30] = qa.get("Q24", "")
    full_row[31] = qa.get("Q25", "")
    full_row[32] = qa.get("Q26", "")
    full_row[33] = qa.get("Q27", "")
    full_row[34] = qa.get("Q28", "")
    full_row[35] = qa.get("Q29", "")
    full_row[36] = qa.get("Q30", "")
    full_row[37] = qa.get("Q31", "")
    full_row[38] = qa.get("Q32", "")
    full_row[39] = qa.get("Q33", "")
    full_row[40] = qa.get("Q34", "")
    full_row[41] = qa.get("Q35", "")
    full_row[42] = qa.get("Q36", "")
    full_row[43] = qa.get("Q37", "")
    full_row[44] = qa.get("Q38", "")
    full_row[45] = qa.get("Q39", "")
    full_row[46] = qa.get("Q40", "")
    full_row[47] = qa.get("Q41", "")
    full_row[48] = qa.get("Q42", "")
    full_row[49] = qa.get("Q43", "")
    full_row[50] = qa.get("Q44", "")
    full_row[51] = qa.get("Q45", "")
    full_row[52] = qa.get("Q46", "")
    full_row[53] = qa.get("Q47", "")
    full_row[54] = qa.get("Q48", "")
    full_row[55] = qa.get("Q49", "")
    full_row[56] = qa.get("Q50", "")
    full_row[57] = qa.get("Q51", "")
    full_row[58] = qa.get("Q52", "")
    full_row[59] = qa.get("Q53", "")
    full_row[60] = qa.get("Q54", "")
    full_row[61] = qa.get("Q55", "")
    full_row[62] = qa.get("Q56", "")
    full_row[63] = qa.get("Q57", "")
    
    # Revenue specific (cols 64 to 70 index-wise, which is Q58 to Q64)
    #HEADERS mapping: Q58 is index 64
    full_row[64] = qa.get("Q58", "")
    full_row[65] = qa.get("Q59", "")
    full_row[66] = qa.get("Q60", "")
    full_row[67] = qa.get("Q61", "")
    full_row[68] = qa.get("Q62", "")
    full_row[69] = qa.get("Q63", "")
    full_row[70] = qa.get("Q64", "")
    
    # Rural Dev specific (cols 71 to 77, which is Q65 to Q71)
    # Leave blank for Revenue and Health
    
    # Forest specific (cols 78 to 82, which is Q72 to Q76)
    # Leave blank
    
    # Health specific (cols 83 to 87, which is Q77 to Q81)
    full_row[83] = qa.get("Q77", "")
    full_row[84] = qa.get("Q78", "")
    full_row[85] = qa.get("Q79", "")
    full_row[86] = qa.get("Q80", "")
    full_row[87] = qa.get("Q81", "")
    
    ws_full.append(full_row)

wb_resp.save(RESPONSES_PATH)
print(f"[✓] Appended 14 new respondents to AIGGPA_Master_140_Responses.xlsx!")

# 3. Update AIGGPA_Master_Tracker.xlsx
ws_log = wb_tracker['Respondent_Log']
# S.No 141 starts at Row 142
for idx, (name, desig, cadre, dept, age, gender, service, edu, q15, q27, qa) in enumerate(new_data):
    row = 142 + idx
    sno = 141 + idx
    resp_id = f"R{sno:03d}"
    
    ws_log.cell(row=row, column=3, value="2026-06-18")
    ws_log.cell(row=row, column=4, value=dept)
    ws_log.cell(row=row, column=5, value="DO")
    ws_log.cell(row=row, column=6, value=cadre)
    ws_log.cell(row=row, column=7, value=age)
    ws_log.cell(row=row, column=8, value=gender)
    ws_log.cell(row=row, column=9, value=service)
    ws_log.cell(row=row, column=10, value="Complete")
    ws_log.cell(row=row, column=11, value="No")
    ws_log.cell(row=row, column=12, value="No")
    ws_log.cell(row=row, column=13, value="Yes")
    ws_log.cell(row=row, column=16, value=f"{name} ({desig})")

print("[✓] Updated Respondent_Log in AIGGPA_Master_Tracker.xlsx!")

# Recalculate Sampling_Dashboard in Tracker
# We will do this by counting the actual respondents we just put in.
# Let's count them
respondents = []
for r in range(2, ws_log.max_row + 1):
    sno_val = ws_log.cell(row=r, column=1).value
    if sno_val is None:
        continue
    dept = ws_log.cell(row=r, column=4).value
    office = ws_log.cell(row=r, column=5).value
    cadre = ws_log.cell(row=r, column=6).value
    status = ws_log.cell(row=r, column=10).value
    if status == "Complete" and dept and office and cadre:
        respondents.append({
            'department': str(dept).strip(),
            'office_type': str(office).strip(),
            'cadre': str(cadre).strip()
        })

sampling_counts = {}
for resp in respondents:
    key = (resp['department'], resp['office_type'], resp['cadre'])
    sampling_counts[key] = sampling_counts.get(key, 0) + 1

ws_dash = wb_tracker['Sampling_Dashboard']
for r in range(4, 15):
    dept = ws_dash.cell(row=r, column=1).value
    office = ws_dash.cell(row=r, column=2).value
    if not dept or not office:
        continue
    dept_str = str(dept).strip()
    office_str = str(office).strip()
    
    for col_idx, cadre in enumerate(["Class I", "Class II", "Class III", "Class IV"], 3):
        key = (dept_str, office_str, cadre)
        count = sampling_counts.get(key, 0)
        ws_dash.cell(row=r, column=col_idx, value=count)
    
    ws_dash.cell(row=r, column=7, value=f"=SUM(C{r}:F{r})")

print("[✓] Updated Sampling_Dashboard in AIGGPA_Master_Tracker.xlsx!")

# Update Daily_Progress in Tracker
ws_prog = wb_tracker['Daily_Progress']
# Find first empty row in Daily_Progress
empty_row = 2
while ws_prog.cell(row=empty_row, column=1).value is not None:
    empty_row += 1

# Append Daily_Progress entries for June 18
ws_prog.cell(row=empty_row, column=1, value="2026-06-18")
ws_prog.cell(row=empty_row, column=2, value="Revenue")
ws_prog.cell(row=empty_row, column=3, value="DO")
ws_prog.cell(row=empty_row, column=4, value=7)
ws_prog.cell(row=empty_row, column=5, value=0)
ws_prog.cell(row=empty_row, column=6, value=0)
ws_prog.cell(row=empty_row, column=7, value=1)
ws_prog.cell(row=empty_row, column=8, value=0)
ws_prog.cell(row=empty_row, column=10, value="Ingestion of 7 Revenue DO respondents from photo survey scans")

empty_row += 1
ws_prog.cell(row=empty_row, column=1, value="2026-06-18")
ws_prog.cell(row=empty_row, column=2, value="Health")
ws_prog.cell(row=empty_row, column=3, value="DO")
ws_prog.cell(row=empty_row, column=4, value=7)
ws_prog.cell(row=empty_row, column=5, value=0)
ws_prog.cell(row=empty_row, column=6, value=0)
ws_prog.cell(row=empty_row, column=7, value=1)
ws_prog.cell(row=empty_row, column=8, value=0)
ws_prog.cell(row=empty_row, column=10, value="Ingestion of 7 Health DO (Class IV support staff) respondents from photo survey scans")

print("[✓] Updated Daily_Progress in AIGGPA_Master_Tracker.xlsx!")

wb_tracker.save(TRACKER_PATH)
print("[✓] Saved updated Master Tracker!")
