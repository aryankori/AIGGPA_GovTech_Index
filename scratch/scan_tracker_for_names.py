import openpyxl
import os

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\08_Data_Entry\Raw_Excel\AIGGPA_Master_DataEntry.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheets:", wb.sheetnames)

for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"Scanning sheet: {sheet}")
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = str(ws.cell(row=r, column=c).value)
            if "Vijendra" in val or "Santere" in val or "Yadav" in val or "Varsha" in val:
                print(f"Found in {sheet} row {r}, col {c}: {val}")

