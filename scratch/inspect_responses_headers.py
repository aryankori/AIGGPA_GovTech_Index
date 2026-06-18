import openpyxl

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\08_Data_Entry\Cleaned_Data\AIGGPA_Master_140_Responses.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheets in Responses:", wb.sheetnames)

ws_resp = wb['Respondent List']
ws_full = wb['Full Responses']

print(f"Respondent List: max_row={ws_resp.max_row}, max_col={ws_resp.max_column}")
print(f"Full Responses: max_row={ws_full.max_row}, max_col={ws_full.max_column}")

# print row 2 of Respondent List
row2_resp = [ws_resp.cell(row=2, column=c).value for c in range(1, ws_resp.max_column + 1)]
print("Respondent List Row 2:", row2_resp)

# print headers of Full Responses
headers_full = [ws_full.cell(row=1, column=c).value for c in range(1, ws_full.max_column + 1)]
print("Full Responses Headers (first 10):", headers_full[:10])
