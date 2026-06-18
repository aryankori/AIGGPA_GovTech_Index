import openpyxl

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\08_Data_Entry\Raw_Excel\AIGGPA_Master_DataEntry.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Data_Matrix']
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print("Headers length:", len(headers))
for idx, h in enumerate(headers, 1):
    print(f"Col {idx}: {h}")
