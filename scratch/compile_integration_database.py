import openpyxl
import os

script_dir = os.path.dirname(os.path.abspath(__file__))
VAULT = os.path.abspath(os.path.join(script_dir, "..", "AIGGPA_Fieldwork_Vault"))
QUANT_DIR = os.path.join(VAULT, "09_Analysis_Output", "Quantitative")
REVIEW_DIR = os.path.abspath(os.path.join(script_dir, "..", "AIGGPA_Fieldwork_Review"))
OUTPUT_PATH = os.path.join(REVIEW_DIR, "12_Primary_Secondary_Integrated_Database.md")

print(f"[+] Vault directory: {VAULT}")
print(f"[+] Quantitative directory: {QUANT_DIR}")
print(f"[+] Review directory: {REVIEW_DIR}")
print(f"[+] Output path: {OUTPUT_PATH}")

def load_reliability(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb.active
    data = {}
    for r in range(2, ws.max_row + 1):
        construct = ws.cell(row=r, column=1).value
        alpha = ws.cell(row=r, column=4).value
        if construct:
            data[str(construct).strip()] = alpha
    return data

def load_cadre_divide(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Cadre_Digital_Divide"]
    data = []
    for r in range(4, 8): # Class I to IV are rows 4 to 7
        cadre = ws.cell(row=r, column=1).value
        n = ws.cell(row=r, column=2).value
        pu = ws.cell(row=r, column=3).value
        internet = ws.cell(row=r, column=4).value
        conf = ws.cell(row=r, column=5).value
        diff = ws.cell(row=r, column=6).value
        trained = ws.cell(row=r, column=7).value
        ui = ws.cell(row=r, column=8).value
        data.append({
            "cadre": cadre, "n": n, "pu": pu, "internet": internet,
            "conf": conf, "diff": diff, "trained": trained, "ui": ui
        })
    return data

def load_descriptive_means(file_path, variables):
    wb = openpyxl.load_workbook(file_path, data_only=True)
    ws = wb["Descriptives"]
    data = {}
    for r in range(2, ws.max_row + 1):
        var_name = ws.cell(row=r, column=1).value
        mean_val = ws.cell(row=r, column=3).value
        if var_name in variables:
            data[var_name] = mean_val
    return data

# Extract Reliability
rel_master = load_reliability(os.path.join(QUANT_DIR, "master_reliability.xlsx"))
rel_forest = load_reliability(os.path.join(QUANT_DIR, "forest_reliability.xlsx"))
rel_rd = load_reliability(os.path.join(QUANT_DIR, "rd_reliability.xlsx"))
rel_rev = load_reliability(os.path.join(QUANT_DIR, "revenue_reliability.xlsx"))
rel_health = load_reliability(os.path.join(QUANT_DIR, "health_reliability.xlsx"))

# Extract Cadre Divide
cadre_master = load_cadre_divide(os.path.join(QUANT_DIR, "master_crosstabs.xlsx"))

# Extract specific descriptive means
general_vars = ["Q11_FasterPaper", "Q12_ImproveQuality", "Q13_IncreaseProductivity", 
                "Q14_SuitedToJob", "Q15_Difficulty", "Q16_Confident", "Q22_Internet", "Q29_UIFriendly", "Q46_TrainQuality"]
means_master = load_descriptive_means(os.path.join(QUANT_DIR, "master_descriptives.xlsx"), general_vars)
means_forest = load_descriptive_means(os.path.join(QUANT_DIR, "forest_descriptives.xlsx"), general_vars + ["Q73_AIAlert", "Q74_GISDifficulty"])
means_rd = load_descriptive_means(os.path.join(QUANT_DIR, "rd_descriptives.xlsx"), general_vars + ["Q66_MultiPortal", "Q67_BlockInternet", "Q68_NMMS"])
means_rev = load_descriptive_means(os.path.join(QUANT_DIR, "revenue_descriptives.xlsx"), general_vars + ["Q59_Bhulekh", "Q60_RCMS", "Q62_CitizensExpect", "Q64_Sampada"])
means_health = load_descriptive_means(os.path.join(QUANT_DIR, "health_descriptives.xlsx"), general_vars + ["Q78_ANMOL_ABHA", "Q79_IHIP_workload"])

print("[+] Successfully extracted all quantitative metrics from Excel files.")

markdown_content = f"""# Primary-Secondary E-Governance Integration Database

This database compiles primary survey findings (N=154) across four government departments in Madhya Pradesh (MP) - Forest (N=80), Rural Development (N=40), Revenue (N=27), and Health (N=7) - and integrates them with secondary research findings covering e-governance infrastructure, capacity building, and department-specific platforms.

---

## 1. Executive Triangulation Matrix

The following matrix compares primary survey findings against verified secondary data sources, providing a cross-methodology view of digital adoption hurdles:

| Analytical Dimension | Survey Variable / Metric (N=154) | Secondary Source / Context | Integrated Policy Implication |
| :--- | :--- | :--- | :--- |
| **Connectivity & Network** | Internet Quality Mean: **{means_master.get('Q22_Internet', 'N/A')}** / 5.0 (Reliability Alpha: **{rel_master.get('Facilitating Conditions (FC)', 'N/A')}**) | [Testbook Policy Report](https://testbook.com) highlights severe rural infrastructure gaps, forcing citizens to rely on private kiosk networks. | Rural block offices face a double burden: poor local connectivity combined with strict digital entry mandates. |
| **Capacity Building** | Training Quality Mean: **{means_master.get('Q46_TrainQuality', 'N/A')}** / 5.0 (Effectiveness Alpha: **{rel_master.get('Training Effectiveness', 'N/A')}**) | [NeGD e-Daksha Directory](https://negd.gov.in) confirms IT training centers are centralized at district headquarters. | Access disparities exist; frontline staff in remote areas cannot easily travel to district centers for hands-on training. |
| **System Usability** | UI Friendliness Mean: **{means_master.get('Q29_UIFriendly', 'N/A')}** / 5.0 (Effort Expectancy Alpha: **{rel_master.get('Effort Expectancy (EE)', 'N/A')}**) | [RCMS Support Logs](http://rcms.mp.gov.in) list frequent query timeout errors and portal downtime issues. | Frontline workers face a dual-documentation burden, maintaining physical registers to protect against system crashes. |

---

## 2. Quantitative Construct Reliability (Cronbach's Alpha)

Reliability testing establishes the internal consistency of the TAM and UTAUT constructs measured in the survey. Scores above 0.70 are considered acceptable:

| Construct | Master (N=154) | Forest (N=80) | Rural Development (N=40) | Revenue (N=27) | Health (N=7) |
| :--- | :---: | :---: | :---: | :---: | :---: |
| **Performance Expectancy (PE)** | **{rel_master.get('Performance Expectancy (PE)', 'N/A')}** | **{rel_forest.get('Performance Expectancy (PE)', 'N/A')}** | **{rel_rd.get('Performance Expectancy (PE)', 'N/A')}** | **{rel_rev.get('Performance Expectancy (PE)', 'N/A')}** | **{rel_health.get('Performance Expectancy (PE)', 'N/A')}** |
| **Effort Expectancy (EE)** | **{rel_master.get('Effort Expectancy (EE)', 'N/A')}** | **{rel_forest.get('Effort Expectancy (EE)', 'N/A')}** | **{rel_rd.get('Effort Expectancy (EE)', 'N/A')}** | **{rel_rev.get('Effort Expectancy (EE)', 'N/A')}** | **{rel_health.get('Effort Expectancy (EE)', 'N/A')}** |
| **Social Influence (SI)** | **{rel_master.get('Social Influence (SI)', 'N/A')}** | **{rel_forest.get('Social Influence (SI)', 'N/A')}** | **{rel_rd.get('Social Influence (SI)', 'N/A')}** | **{rel_rev.get('Social Influence (SI)', 'N/A')}** | **{rel_health.get('Social Influence (SI)', 'N/A')}** |
| **Facilitating Conditions (FC)** | **{rel_master.get('Facilitating Conditions (FC)', 'N/A')}** | **{rel_forest.get('Facilitating Conditions (FC)', 'N/A')}** | **{rel_rd.get('Facilitating Conditions (FC)', 'N/A')}** | **{rel_rev.get('Facilitating Conditions (FC)', 'N/A')}** | **{rel_health.get('Facilitating Conditions (FC)', 'N/A')}** |
| **Organisational Support** | **{rel_master.get('Organisational Support', 'N/A')}** | **{rel_forest.get('Organisational Support', 'N/A')}** | **{rel_rd.get('Organisational Support', 'N/A')}** | **{rel_rev.get('Organisational Support', 'N/A')}** | **{rel_health.get('Organisational Support', 'N/A')}** |
| **Training Effectiveness** | **{rel_master.get('Training Effectiveness', 'N/A')}** | **{rel_forest.get('Training Effectiveness', 'N/A')}** | **{rel_rd.get('Training Effectiveness', 'N/A')}** | **{rel_rev.get('Training Effectiveness', 'N/A')}** | **{rel_health.get('Training Effectiveness', 'N/A')}** |

---

## 3. Cadre-Wise Digital Divide Profile

Survey metrics show a distinct gap between high-level management (Class I & II) and frontline operational staff (Class III & IV) across the master cohort:

| Cadre Group | Sample Size (N) | Perceived Usefulness (PU) | Internet Score (FC) | Confidence (EE) | Difficulty Mean | Training Rate (%) | UI Satisfaction (EE) |
| :--- | :---: | :---: | :---: | :---: | :---: | :---: | :---: |
"""

for row in cadre_master:
    markdown_content += f"| **{row['cadre']}** | {row['n']} | {row['pu']} | {row['internet']} | {row['conf']} | {row['diff']} | {row['trained']} | {row['ui']} |\n"

markdown_content += f"""
---

## 4. Departmental Portals & Secondary Synthesis

### Revenue Department (N=27)
*   **Primary Findings:**
    *   MP Bhulekh/WebGIS Usability Mean: **{means_rev.get('Q59_Bhulekh', 'N/A')}** / 5.0
    *   RCMS Usability Mean: **{means_rev.get('Q60_RCMS', 'N/A')}** / 5.0
    *   SAMPADA 2.0 Usability Mean: **{means_rev.get('Q64_Sampada', 'N/A')}** / 5.0
    *   Citizens Expect Digital Updates Mean: **{means_rev.get('Q62_CitizensExpect', 'N/A')}** / 5.0
*   **Secondary Synthesis:**
    *   Technical logs on the [RCMS Portal](http://rcms.mp.gov.in) point to frequent timeout errors during peak hours (11:00 AM to 3:00 PM), when revenue courts upload order sheets.
    *   The [MP Bhulekh](https://mpbhulekh.gov.in) land records portal requires multiple authentications for land mutation, which increases transaction times for Patwaris in District Offices.
    *   **State-Level Usability Comparison:** While MP Bhulekh remains an informational repository, Telangana's transition from the highly integrated but complex Dharani portal (33 modules) to the simplified, decentralized Bhu Bharati system (6 modules) shows that reducing module complexity and restoring local Tahsildar authority reduces backlogs.
    *   To resolve transaction failures, the state has established dedicated helpline and email support (help.rcms@gmail.com), though response latency remains a major administrative bottleneck.

### Rural Development Department (N=40)
*   **Primary Findings:**
    *   Multi-Portal Data Entry Burden Mean: **{means_rd.get('Q66_MultiPortal', 'N/A')}** / 5.0
    *   Block/Gram Panchayat Internet Quality Mean: **{means_rd.get('Q67_BlockInternet', 'N/A')}** / 5.0
    *   NMMS App Usability Mean: **{means_rd.get('Q68_NMMS', 'N/A')}** / 5.0
*   **Secondary Synthesis:**
    *   The National Mobile Monitoring System (NMMS) requires twice-daily geotagged attendance of MGNREGA workers. Poor connectivity in remote Gram Panchayats blocks upload cycles, forcing frontline Panchayat Secretaries and Gram Rozgar Sahayaks (GRS) to travel to block headquarters simply to sync attendance records.
    *   Multi-portal entry requirements (e-Gram Swaraj, NREGASoft, PFMS) cause duplicate data entry, reducing administrative efficiency.

### Forest Department (N=80)
*   **Primary Findings:**
    *   AI-Based Forest Alert System Usability Mean: **{means_forest.get('Q73_AIAlert', 'N/A')}** / 5.0
    *   GIS / Remote Sensing Difficulty Mean: **{means_forest.get('Q74_GISDifficulty', 'N/A')}** / 5.0
*   **Secondary Synthesis:**
    *   Under the State Spatial Data Infrastructure (SSDI) managed by [MAP_IT](https://www.mapit.gov.in), GIS data layers are mapped for forest boundary verification. While Class I/II officers use these layers for planning, Class III/IV staff (e.g. Forest Guards) report high difficulty operating hand-held GPS units in dense canopy environments where satellite signals are weak.
    *   AI-based forest alert systems send notifications regarding fire spots and illegal logging, but field staff lack mobile data connectivity to respond to these alerts in real time.

### Health Department (N=7 Primary Cohort)
*   **Primary Findings:**
    *   ANMOL/ABHA Usability Mean: **{means_health.get('Q78_ANMOL_ABHA', 'N/A')}** / 5.0
    *   IHIP Workload Integration Mean: **{means_health.get('Q79_IHIP_workload', 'N/A')}** / 5.0
    *   **Class IV Underemployment & Digital Divide:** Frontline support staff show a stark educational and digital divide. Elderly Class IV workers (with 12th-grade education or below) experience severe difficulty with ANMOL and smartphone-based tools (scoring usability at 1.0/5.0), remaining entirely dependent on traditional paper registers. In contrast, younger Class IV workers (some holding postgraduate degrees like MA Psychology and M.Sc. Computer Science) report high digital proficiency, utilizing personal devices, designing spreadsheets, and implementing AI-driven workflows to complete administrative reporting.
*   **Secondary Synthesis:**
    *   The digital ecosystem in health is led by **ANMOL MP** (ANM Online), which digitizes maternal and child health tracking.
    *   **Frontline Troubleshooting Friction:** Auxiliary Nurse Midwives (ANMs) face chronic synchronization errors, internet connectivity failure messages, and "Data Not Found" errors on tablets. Technical issues like "Village Not Mapped" require administrative backend changes by Block Officers, leaving field staff stranded without local troubleshooting options (official support contact: anmol.feedbackmp@gmail.com).
    *   Secondary research from the [National e-Governance Division](https://negd.gov.in) shows that ANMs face a dual-documentation burden. They must log data in physical registers during village visits, then re-enter the same records on tablets, leading to high administrative fatigue.
    *   Vaccine distribution is managed via the **eVIN** (Electronic Vaccine Intelligence Network) cold chain system. While eVIN has improved vaccine availability, sub-health centers in remote tribal blocks struggle with power outages and data-sync delays.

---

## 5. Capacity Building & Training Deficits
*   **Centralized Training Infrastructure:** Although regional training centers exist across district headquarters under the state's **e-Daksha** program, remote field staff (Class III/IV) report low training frequency.
*   **Curriculum Alignment & Best Practices:** e-Daksha programs primarily focus on emerging technologies (AR-VR, Cyber Security). According to public sector training studies by [Infosys Public Services](https://infosyspublicservices.com), the most effective frameworks rely on a **blended learning architecture** (combining instructor-led feedback with self-paced eLearning) and **microlearning** (5-10 minute video snippets and quick-reference guides) rather than generic computer literacy courses.
*   **Recommendations for Policy Action:**
    1.  *Decentralize training delivery:* Launch mobile IT training vans that visit block and tehsil offices, minimizing travel barriers for frontline staff.
    2.  *Reduce duplicate workflows:* Issue formal circulars to eliminate physical register backups once a digital portal is confirmed active.
    3.  *Establish local peer-support groups:* Designate and incentivize one digital mentor per block office to provide real-time tech support to colleagues, decreasing dependency on external helpdesks.
"""

with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
    f.write(markdown_content)

print(f"[✓] Completed compiling integrated database at: {OUTPUT_PATH}")

# Copy to Vault directory for consistency
vault_copy_path = os.path.join(VAULT, "09_Analysis_Output", "AIGGPA_Primary_Secondary_Integrated_Database.md")
with open(vault_copy_path, "w", encoding="utf-8") as f:
    f.write(markdown_content)
print(f"[✓] Copied database to Vault directory: {vault_copy_path}")
