import openpyxl

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\08_Data_Entry\Raw_Excel\AIGGPA_Master_DataEntry.xlsx"
wb = openpyxl.load_workbook(path, data_only=False) # Load formulas instead of values
ws = wb['Data_Matrix']
for c in range(79, 83):
    val = ws.cell(row=2, column=c).value
    header = ws.cell(row=1, column=c).value
    print(f"Col {c} ({header}): {val}")
