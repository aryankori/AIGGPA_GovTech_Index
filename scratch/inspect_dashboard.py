import openpyxl

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\AIGGPA_Master_Tracker.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Sampling_Dashboard']
for r in range(1, ws.max_row + 1):
    vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
    print(f"Row {r}: {vals}")
