import openpyxl
import os

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\08_Data_Entry\Raw_Excel\AIGGPA_Master_DataEntry.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
print("Sheets:", wb.sheetnames)

for name in wb.sheetnames:
    ws = wb[name]
    print(f"\nSheet {name}: max_row={ws.max_row}, max_column={ws.max_column}")
    # print headers
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print("Headers count:", len(headers))
    print("Headers:", headers[:10])
