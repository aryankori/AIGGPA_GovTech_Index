import openpyxl

path = r"c:\Users\aryan\OneDrive\Documents\Visual Studio 2022\AIGGPA_Report\AIGGPA_Fieldwork_Vault\08_Data_Entry\Cleaned_Data\AIGGPA_Master_140_Responses.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb['Full Responses']
for c in range(75, ws.max_column + 1):
    print(f"Col {c}: {ws.cell(row=1, column=c).value}")
