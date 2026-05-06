import openpyxl
from openpyxl.styles import Alignment, PatternFill
import copy

def copy_format():
    # Load our current data
    data_wb = openpyxl.load_workbook(r"C:\Users\aryan\Downloads\Framework for Interns.xlsx")
    data_ws = data_wb.active
    
    # Load the original boss template
    template_wb = openpyxl.load_workbook(r"C:\Users\aryan\Downloads\Framework for Interns (1).xlsx")
    template_ws = template_wb.active
    
    # Unmerge all cells in the template
    for merged_cell in list(template_ws.merged_cells.ranges):
        template_ws.unmerge_cells(str(merged_cell))
        
    # Copy header values if any exist from our data
    template_ws["B1"] = data_ws["B1"].value
    
    # Extract fill colors explicitly from template proxy
    fill1 = copy.copy(template_ws["A4"].fill)
    fill2 = copy.copy(template_ws["A8"].fill)
    fill3 = copy.copy(template_ws["A12"].fill)
    fill4 = PatternFill(start_color='FFC9DAF8', end_color='FFC9DAF8', fill_type='solid')
    
    colors = [fill1, fill2, fill3, fill4]
    
    for i in range(4):
        src_row = i + 3
        dest_row = i + 4
        
        for col in range(1, 13): # A to L
            val = data_ws.cell(row=src_row, column=col).value
            target_cell = template_ws.cell(row=dest_row, column=col)
            
            target_cell.value = val
            target_cell.fill = copy.copy(colors[i])
            target_cell.alignment = Alignment(wrap_text=True, vertical='top')
            
    # Adjust column widths to make it readable like before
    for col in template_ws.columns:
        col_letter = col[0].column_letter
        if col_letter in ['A', 'B', 'C', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
            template_ws.column_dimensions[col_letter].width = 35
            
    # clear row 8 to 15 just in case there's leftover text from template
    for row in range(8, 20):
        for col in range(1, 13):
            template_ws.cell(row=row, column=col).value = None
            template_ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)
            
    template_wb.save(r"C:\Users\aryan\Downloads\Framework for Interns.xlsx")
    print("Formatting restored from template with condensed layout.")

if __name__ == "__main__":
    copy_format()
