import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os

def build_expanded_framework():
    filepath = r"C:\Users\aryan\Downloads\Framework for Interns.xlsx"
    wb = openpyxl.Workbook() # create a fresh one to avoid merge conflicts
    ws = wb.active
    ws.title = "Framework"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1B2A4A") # Navy blue
    bold_font = Font(bold=True)
    wrap_align = Alignment(wrap_text=True, vertical="center", horizontal="left")
    center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")
    
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                         top=Side(style='thin'), bottom=Side(style='thin'))
                         
    # Row 1: Title
    ws.merge_cells("A1:L1")
    ws["A1"] = "Title of the Research/Study: Assessment of Digital Tool Adoption and Its Impact on Efficiency in MP Government Departments (Revenue, Rural Development, Forest, Health)"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = center_align
    ws["A1"].fill = PatternFill("solid", fgColor="D6E4F0")
    
    # Headers
    headers = [
        "Research Objectives", "Research Questions", "Indicators", 
        "Quantitative / Qualitative", "Analysis Tools", "Visual Representation", 
        "Levels", "Schedule/Questionnaire Questions", "Means of Verification", 
        "Findings", "Doubts", "References"
    ]
    
    for col_num, header_title in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header_title
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Data structure: each objective is a dict with meta (merged) and rows (unmerged)
    # Colors: RO1 (Yellow), RO2 (Pink/Purple), RO3 (Green), RO4 (Blue)
    data = [
        {
            "color": "FFF2CC", # Light yellow
            "obj": "RO1: Evaluate the current status of digital infrastructure and employee awareness of tools.",
            "rq": "What is the current status of digital infrastructure in the selected departments, and how aware are employees of the tools available to them?",
            "verification": "Direct office observation.\nChecking department IT inventories.\nComparing survey answers with what we actually see.",
            "findings": "(To be filled after fieldwork)",
            "doubts": "How should we count shared devices? Per device or per user?",
            "refs": "Davis (1989) - TAM\nHeeks (2006)",
            "rows": [
                [
                    "Number and type of devices per employee", "Quantitative", 
                    "Descriptive statistics (means, frequencies)", "Bar charts for device availability", 
                    "Head Office and District Office", "Q1. What devices are at your workstation?"
                ],
                [
                    "Internet bandwidth and uptime", "Quantitative", 
                    "Cross-tabbing by department and cadre", "Likert distribution charts", 
                    "All 4 departments", "Q2. Rate your internet connectivity (1-5)."
                ],
                [
                    "Awareness percentage for key portals", "Quantitative", 
                    "Percentage / Proportion Analysis", "Stacked bars for awareness by age", 
                    "All 4 departments", "Q3. Which of these tools do you know about?\n[Revenue: RCMS, SAMPADA]\n[Forest: e-Green Watch]"
                ],
                [
                    "Daily usage frequency & specific tools used", "Mixed", 
                    "Mapping to TAM 'Facilitating Conditions'", "Heatmaps showing tool usage across cadres", 
                    "All 4 departments", "Q4. How often do you use them?\nQ5. (Observation) Physical setup notes."
                ]
            ]
        },
        {
            "color": "F4CCCC", # Light pink
            "obj": "RO2: Assess capacity building efforts and the quality of technical training provided.",
            "rq": "What training and technical support do departments provide to help staff use digital tools?",
            "verification": "Official training circulars.\niGOT/Mission Karmayogi enrollment data.\nCross-checking staff claims against official records.",
            "findings": "(To be filled after fieldwork)",
            "doubts": "Departments might not keep good training records. How do we verify attendance if HR doesn't have the data?",
            "refs": "Venkatesh et al. (2003) - UTAUT\nGovt of India (2020)",
            "rows": [
                [
                    "Training programs held in the last 2 years & % of staff who attended", "Quantitative", 
                    "Descriptive stats & Frequencies", "Grouped bar charts for training by department", 
                    "Head Office and District Office", "Q6. Have you received formal IT training?\nQ7. How many sessions in the last 2 years?"
                ],
                [
                    "Staff satisfaction with training quality", "Quantitative", 
                    "Likert-scale averages", "Radar charts for training quality metrics", 
                    "All 4 departments", "Q8. Rate the quality and relevance (1-5)."
                ],
                [
                    "Availability of an IT helpdesk & subjective feedback", "Qualitative", 
                    "Thematic analysis for interview transcripts", "Word clouds of training gaps", 
                    "All 4 departments", "Q9. Is there an IT helpdesk here?\nQ10. (Interview) What was useful about the training, and what was missing?"
                ]
            ]
        },
        {
            "color": "D9EAD3", # Light green
            "obj": "RO3: Identify systemic bottlenecks, technical issues, and attitudinal barriers.",
            "rq": "What are the main roadblocks employees face when using digital tools in their daily work?",
            "verification": "Comparing survey, interview, and FGD data to find common patterns.\nObserving actual disruptions during visits.",
            "findings": "(To be filled after fieldwork)",
            "doubts": "Staff might underreport issues because they fear backlash from management. We need strict anonymity protocols.",
            "refs": "Braun & Clarke (2006)\nUnited Nations (2024)",
            "rows": [
                [
                    "Common technical issues (internet, crashes) & system downtime", "Quantitative", 
                    "Frequency counts for reported issues", "Pareto charts for top bottlenecks", 
                    "Head Office and District Office", "Q11. What are your biggest challenges with digital tools?\nQ12. How often do you face technical issues?"
                ],
                [
                    "Perceived difficulty of the tools (Effort Expectancy)", "Quantitative", 
                    "Ranking analysis to prioritize problems", "Diverging bar charts for difficulty ratings", 
                    "All Cadres (I to IV)", "Q13. How difficult are the tools to use (1-5)?\n[Forest: GIS tools]\n[Rural Dev: Multi-portal load]"
                ],
                [
                    "Adequacy of support when things break", "Quantitative", 
                    "Cross-tabulation by Cadre", "Stacked bar charts", 
                    "All Cadres (I to IV)", "Q14. Do you get enough support when things break?"
                ],
                [
                    "Attitudinal & Organizational barriers (Systemic)", "Qualitative", 
                    "Thematic analysis (main approach)", "Concept maps from interview transcripts", 
                    "All Cadres (I to IV)", "Q15. (FGD) What systemic changes would make things easier?"
                ]
            ]
        },
        {
            "color": "C9DAF8", # Light blue
            "obj": "RO4: Develop actionable recommendations to improve digital efficiency.",
            "rq": "What specific steps can we suggest to improve digital tool usage and public service delivery?",
            "verification": "Ensuring all recommendations are backed by hard data from Objectives 1-3.\nChecking feasibility against known budget constraints.",
            "findings": "(To be filled after fieldwork)",
            "doubts": "Since this relies on the other objectives, we can't finalize it until fieldwork is done. Should we draft preliminary ideas based on literature first?",
            "refs": "OECD Digital Govt Framework (2014)",
            "rows": [
                [
                    "Priority improvement areas ranked by staff", "Quantitative", 
                    "Priority matrix mapping", "Impact vs. Feasibility priority matrix", 
                    "Department-level interventions", "Q16. What specific improvements do you need?\nQ17. If you could pick one area to fix, what would it be?"
                ],
                [
                    "Specific software/hardware requests", "Qualitative", 
                    "Gap analysis (current vs. desired state)", "Summary tables for recommendations", 
                    "Department-level interventions", "Q18. (Interview) In an ideal world, how would digital tools support your job?"
                ],
                [
                    "Proposed policy changes", "Qualitative", 
                    "Comparing findings against standard benchmarks", "Timeline roadmap for short/medium/long-term actions", 
                    "State-level policy recommendations", "Q19. (FGD) What policy changes are most urgent?"
                ]
            ]
        }
    ]
    
    current_row = 3
    for block in data:
        start_row = current_row
        num_rows = len(block["rows"])
        end_row = current_row + num_rows - 1
        
        fill_color = PatternFill("solid", fgColor=block["color"])
        
        for i, row_data in enumerate(block["rows"]):
            # C through H
            for j, val in enumerate(row_data):
                cell = ws.cell(row=current_row+i, column=j+3)
                cell.value = val
                cell.alignment = wrap_align
                cell.fill = fill_color
                cell.border = thin_border
                
        # Merge A, B, I, J, K, L across the rows of this objective
        merge_cols = [
            (1, block["obj"]), (2, block["rq"]), 
            (9, block["verification"]), (10, block["findings"]), 
            (11, block["doubts"]), (12, block["refs"])
        ]
        
        for col_idx, val in merge_cols:
            if num_rows > 1:
                ws.merge_cells(start_row=start_row, start_column=col_idx, end_row=end_row, end_column=col_idx)
            cell = ws.cell(row=start_row, column=col_idx)
            cell.value = val
            cell.alignment = center_align
            cell.font = bold_font if col_idx in [1, 2] else Font()
            cell.fill = fill_color
            cell.border = thin_border
            
            # ensure borders on merged cells
            for r in range(start_row, end_row + 1):
                ws.cell(row=r, column=col_idx).border = thin_border
                
        current_row += num_rows

    # Set column widths
    widths = {
        "A": 25, "B": 25, "C": 30, "D": 15, "E": 25, "F": 25, 
        "G": 20, "H": 30, "I": 25, "J": 20, "K": 25, "L": 20
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
        
    wb.save(filepath)
    print("Framework successfully expanded and color-coded!")

if __name__ == "__main__":
    build_expanded_framework()
