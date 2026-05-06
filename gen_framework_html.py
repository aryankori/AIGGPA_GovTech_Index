"""Read Framework_Completed.xlsx and generate a styled HTML page."""
import openpyxl, html as h

wb = openpyxl.load_workbook('Framework_Completed.xlsx')
ws = wb.active

title = ws['A1'].value
headers = [ws.cell(row=3, column=c).value or '' for c in range(1, 12)]
headers = [hdr.replace('\n', ' ') for hdr in headers]

objs = []
for start_row in [4, 8, 12, 16]:
    row_data = []
    for c in range(1, 12):
        val = ws.cell(row=start_row, column=c).value or ''
        row_data.append(str(val))
    objs.append(row_data)

likert = []
for r in range(5, 10):
    s = ws.cell(row=r, column=13).value
    m = ws.cell(row=r, column=14).value
    if s: likert.append((str(s), str(m)))

obj_colors = ['#D6E4F0', '#DFF0D8', '#FDEBD0', '#E8DAEF']
obj_accent = ['#2E5090', '#2E7D32', '#E65100', '#6A1B9A']
obj_titles = [
    'Digital Infrastructure, Awareness & Usage',
    'Capacity Building & Training Assessment',
    'Bottlenecks, Issues & Challenges',
    'Recommendations for Improvement'
]

css = """
:root{--navy:#1B2A4A;--gold:#C49A2A;--dark:#2D2D2D;--gray:#6B7280;--light:#F8F9FA;--white:#FFFFFF}
*{margin:0;padding:0;box-sizing:border-box}
body{font-family:'Segoe UI',system-ui,-apple-system,sans-serif;background:linear-gradient(135deg,#0f172a 0%,#1e293b 50%,#0f172a 100%);color:var(--dark);min-height:100vh}
.hero{text-align:center;padding:60px 20px 40px;color:var(--white)}
.hero h1{font-size:2.2rem;font-weight:700;letter-spacing:-0.5px;max-width:800px;margin:0 auto 12px}
.hero .divider{width:120px;height:3px;background:var(--gold);margin:0 auto 16px;border-radius:2px}
.hero p{font-size:1rem;color:#94a3b8;max-width:600px;margin:0 auto}
.container{max-width:1200px;margin:0 auto;padding:0 20px 60px}
.obj-card{background:var(--white);border-radius:16px;margin-bottom:32px;overflow:hidden;box-shadow:0 4px 24px rgba(0,0,0,0.15)}
.obj-header{padding:20px 28px;display:flex;align-items:center;gap:16px}
.obj-num{width:48px;height:48px;border-radius:12px;display:flex;align-items:center;justify-content:center;font-size:1.4rem;font-weight:800;color:var(--white);flex-shrink:0}
.obj-header h2{font-size:1.25rem;font-weight:700;color:var(--dark);margin:0}
.obj-header h2 span{display:block;font-size:0.8rem;font-weight:400;color:var(--gray);margin-top:2px}
.obj-body{padding:0 28px 28px}
.field{margin-bottom:20px}
.field-label{font-size:0.7rem;font-weight:700;text-transform:uppercase;letter-spacing:1.2px;color:var(--gray);margin-bottom:6px}
.field-content{font-size:0.92rem;line-height:1.65;color:#374151;white-space:pre-line;background:var(--light);padding:14px 18px;border-radius:10px;border-left:3px solid var(--gold)}
.grid-2{display:grid;grid-template-columns:1fr 1fr;gap:16px}
.grid-3{display:grid;grid-template-columns:1fr 1fr 1fr;gap:16px}
@media(max-width:768px){.grid-2,.grid-3{grid-template-columns:1fr}}
.likert-section{background:var(--white);border-radius:16px;padding:28px;box-shadow:0 4px 24px rgba(0,0,0,0.15);margin-bottom:32px}
.likert-section h2{font-size:1.25rem;color:var(--dark);margin-bottom:16px;display:flex;align-items:center;gap:8px}
.likert-table{width:100%;border-collapse:separate;border-spacing:0;border-radius:10px;overflow:hidden}
.likert-table th{background:var(--navy);color:var(--white);padding:12px 16px;font-size:0.85rem;text-align:left}
.likert-table td{padding:10px 16px;font-size:0.9rem;border-bottom:1px solid #e5e7eb}
.likert-table tr:nth-child(even) td{background:#f9fafb}
.likert-table tr:last-child td{border-bottom:none}
.badge{display:inline-block;padding:3px 10px;border-radius:20px;font-size:0.75rem;font-weight:600}
footer{text-align:center;padding:20px;color:#64748b;font-size:0.8rem}
"""

field_map = [
    (1, 'Research Question'),
    (2, 'Indicators'),
    (3, 'Quantitative / Qualitative'),
    (4, 'Analysis Tools'),
    (5, 'Visual Representation'),
    (6, 'Levels'),
    (7, 'Schedule / Questionnaire Questions'),
    (8, 'Means of Verification'),
    (9, 'Findings'),
    (10, 'Doubts'),
]

cards_html = ''
for i, obj in enumerate(objs):
    fields_html = ''
    pairs = []
    for idx, label in field_map:
        val = h.escape(obj[idx]).replace('\n', '<br>')
        block = f'<div class="field"><div class="field-label">{h.escape(label)}</div><div class="field-content" style="border-left-color:{obj_accent[i]}">{val}</div></div>'
        pairs.append((label, block))

    # Layout: question full width, then pairs in grids
    fields_html += pairs[0][1]  # Research Question full width

    # Indicators + Quant/Qual
    fields_html += f'<div class="grid-2">{pairs[1][1]}{pairs[2][1]}</div>'
    # Analysis + Visual
    fields_html += f'<div class="grid-2">{pairs[3][1]}{pairs[4][1]}</div>'
    # Levels + Questions
    fields_html += f'<div class="grid-2">{pairs[5][1]}{pairs[6][1]}</div>'
    # Verification + Findings + Doubts
    fields_html += f'<div class="grid-3">{pairs[7][1]}{pairs[8][1]}{pairs[9][1]}</div>'

    cards_html += f'''
    <div class="obj-card">
        <div class="obj-header" style="background:{obj_colors[i]}">
            <div class="obj-num" style="background:{obj_accent[i]}">{i+1}</div>
            <h2>{h.escape(obj_titles[i])}<span>Research Objective {i+1}</span></h2>
        </div>
        <div class="obj-body">{fields_html}</div>
    </div>'''

likert_rows = ''
for s, m in likert:
    likert_rows += f'<tr><td><strong>{h.escape(s)}</strong></td><td>{h.escape(m)}</td></tr>'

page = f'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>AIGGPA Intern Research Framework</title>
<style>{css}</style>
</head>
<body>
<div class="hero">
    <h1>{h.escape(title)}</h1>
    <div class="divider"></div>
    <p>Intern Research Framework — AIGGPA, Bhopal | April 2026</p>
</div>
<div class="container">
    {cards_html}
    <div class="likert-section">
        <h2>📊 Likert Scale Reference</h2>
        <table class="likert-table">
            <thead><tr><th>Scale Point</th><th>Meaning</th></tr></thead>
            <tbody>{likert_rows}</tbody>
        </table>
    </div>
</div>
<footer>AIGGPA Research Framework &middot; Prepared by Aryan Kori &middot; April 2026</footer>
</body>
</html>'''

out = 'Framework_View.html'
with open(out, 'w', encoding='utf-8') as f:
    f.write(page)
print(f'Done: {out} ({len(page):,} bytes)')
