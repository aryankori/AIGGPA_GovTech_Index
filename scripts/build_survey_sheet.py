"""
Build two outputs:
1. Updated mp_forest_directory.csv with a global Serial Number column
2. mp_forest_survey.xlsx — a survey response sheet where each row = one person (by serial no.)
   and each column = one questionnaire question, with dropdown validation for options.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
import os

# ── 1. Add serial numbers to the directory CSV ──────────────────────────
csv_path = "mp_forest_directory.csv"
df = pd.read_csv(csv_path, encoding='utf-8-sig')
if 'Serial No. (क्रमांक)' not in df.columns:
    df.insert(0, 'Serial No. (क्रमांक)', range(1, len(df) + 1))
    df.to_csv(csv_path, index=False, encoding='utf-8-sig')
    print(f"✅ Added serial numbers 1–{len(df)} to {csv_path}")
else:
    print(f"✅ Serial numbers already exist in {csv_path} ({len(df)} rows)")

# ── 2. Define ALL 81 questions with their options ───────────────────────
# Format: (question_number, question_text, type, options_list_or_None)
# type: 'text' = free text, 'radio' = single choice, 'check' = multi-select, 'likert' = 1-5 scale, 'rank' = ranking

questions = [
    # Section 1: Personal Information
    (1, "Name / नाम", "prefilled", None),
    (2, "Designation / Post / पदनाम", "prefilled", None),
    (3, "Mobile Number / मोबाइल नंबर", "prefilled", None),
    (4, "Email Address / ईमेल पता", "prefilled", None),
    (5, "Job Role / Level / कार्य भूमिका", "text", None),
    (6, "Age group / आयु वर्ग", "radio", ["Below 30", "30-45", "46-60"]),
    (7, "Gender / लिंग", "radio", ["Male", "Female", "Other"]),
    (8, "Years of service / सेवा के वर्ष", "radio", ["0-5", "6-10", "11-20", "21+"]),
    (9, "Highest education / उच्चतम शिक्षा", "radio", ["Up to 12th", "Grad", "PG", "Prof"]),

    # Section 2: Digital Tools — Opinion Attitude
    (10, "Should govt employees adopt digital tools? / डिजिटल उपकरण अपनाने चाहिए?", "radio", ["Yes, strongly agree", "Yes, somewhat", "Neutral", "No, not necessary", "No, not suitable"]),
    (11, "Digital tools help complete tasks faster than paper / कार्य तेज़ी से पूरा", "likert", ["1","2","3","4","5"]),
    (12, "Digital tools improve quality/accuracy / गुणवत्ता/सटीकता में सुधार", "likert", ["1","2","3","4","5"]),
    (13, "Digital tools increase overall productivity / उत्पादकता बढ़ाता है", "likert", ["1","2","3","4","5"]),
    (14, "Digital tools well-suited to actual job tasks / कार्यों के लिए उपयुक्त", "likert", ["1","2","3","4","5"]),
    (15, "How difficult are digital tools to use? / कितना कठिन", "likert", ["1","2","3","4","5"]),
    (16, "Confident in ability to use digital tools / क्षमता पर विश्वास", "likert", ["1","2","3","4","5"]),
    (17, "Superiors encourage digital tool use / वरिष्ठ प्रोत्साहित करते हैं", "likert", ["1","2","3","4","5"]),
    (18, "Colleagues regularly use digital tools / सहकर्मी नियमित उपयोग", "likert", ["1","2","3","4","5"]),
    (19, "Formal mandate for digital tool use? / औपचारिक आदेश है?", "radio", ["Yes", "No", "Don't know"]),

    # Section 3: IT Infrastructure Connectivity
    (20, "Digital devices at workstation / कार्यस्थल पर उपकरण", "check", ["Desktop", "Laptop", "Tablet", "Phone", "None"]),
    (21, "Share device with others? / उपकरण साझा करते हैं?", "radio", ["Yes always", "Sometimes", "No dedicated"]),
    (22, "Rate internet connectivity / इंटरनेट कनेक्टिविटी मूल्यांकन", "likert", ["1","2","3","4","5"]),
    (23, "Internet outages per week / सप्ताह में इंटरनेट बंद", "radio", ["Never", "1-2", "3-5", "Daily"]),
    (24, "IT helpdesk available? / IT हेल्पडेस्क उपलब्ध?", "radio", ["Yes", "No"]),
    (25, "How quickly issues resolved? / समस्याएँ कितनी जल्दी हल?", "radio", ["Same day", "2-3 days", "1 week+", "Never"]),

    # Section 4: General Digital Tool Usage
    (26, "How often use digital tools? / कितनी बार उपयोग?", "radio", ["Daily", "Weekly", "Monthly", "Rarely", "Never"]),
    (27, "% work done digitally / कितना प्रतिशत डिजिटल?", "radio", ["0-20", "21-40", "41-60", "61-80", "81-100"]),
    (28, "Time to learn new portal/app / नया पोर्टल सीखने में समय", "radio", ["<1 day", "Few days", "1-2 weeks", ">2 weeks"]),
    (29, "Govt portal design user-friendly / डिज़ाइन उपयोगकर्ता-अनुकूल", "likert", ["1","2","3","4","5"]),
    (30, "General tools awareness / सामान्य उपकरणों से अवगत", "check", ["e-Office", "CM Helpline", "PFMS", "SPARROW", "iGOT", "MP eDistrict"]),
    (31, "Primary interaction with digital tools / प्राथमिक भूमिका", "radio", ["Data entry", "Review/approve", "Field verification", "Don't use", "Other"]),
    (32, "One person does portal work for others? / एक व्यक्ति पोर्टल कार्य?", "radio", ["Yes one person", "A few share", "Everyone does own", "N/A"]),
    (33, "Senior officers use digital tools themselves? / वरिष्ठ स्वयं उपयोग?", "radio", ["Yes", "Rely on subordinates", "Mixed", "Don't know"]),
    (34, "Digital tools changed work expected? / अपेक्षित कार्य बदला?", "likert", ["1","2","3","4","5"]),
    (35, "When portal gives error, what do you do? / पोर्टल त्रुटि पर क्या करें?", "radio", ["Wait for IT", "Ask colleague", "Use paper", "Fix myself", "Tell supervisor", "Abandon"]),

    # Section 5: Other Digital Tools
    (36, "Use non-govt apps for work? / गैर-सरकारी ऐप्स उपयोग?", "radio", ["Yes", "No"]),
    (37, "Which tools used for work? / कार्य के लिए कौन-से?", "check", ["WhatsApp", "Google Docs/Drive", "ChatGPT/AI", "YouTube", "Personal email", "MS Office", "Google Translate", "Other"]),
    (38, "What do you use these tools for? / किसलिए उपयोग?", "check", ["Drafting", "Translating", "Coordinating", "Learning portals", "Backup", "Sharing files", "Other"]),
    (39, "How often use personal tools for official work? / कितनी बार?", "radio", ["Daily", "Few times/week", "Occasionally", "Rarely", "Never"]),
    (40, "Personal tools fill gap govt systems don't cover? / कमी पूरी करते हैं?", "likert", ["1","2","3","4","5"]),
    (41, "Concerns about using personal tools for official work? / कोई चिंता?", "text", None),

    # Section 6: Barriers Challenges
    (42, "Issues faced / समस्याएँ", "check", ["Slow internet", "Crashes", "No device", "Complex UI", "No training", "No support", "Power cuts"]),
    (43, "How often digital issues disrupt work? / कितनी बार बाधा?", "radio", ["Daily", "Weekly", "Monthly", "Rarely", "Never"]),

    # Section 7: Training Support Needs
    (44, "Digital skills training in last 2 years? / प्रशिक्षण लिया?", "radio", ["Yes", "No"]),
    (45, "If yes, how many sessions? / कितने सत्र?", "radio", ["1", "2-3", "4-5", "More than 5"]),
    (46, "Rate quality of training / प्रशिक्षण गुणवत्ता", "likert", ["1","2","3","4","5"]),
    (47, "Training sufficient for job needs? / पर्याप्त था?", "likert", ["1","2","3","4","5"]),
    (48, "Topics needing more training / अधिक प्रशिक्षण चाहिए?", "text", None),
    (49, "Training appropriate for job role? / भूमिका के लिए उपयुक्त?", "likert", ["1","2","3","4","5"]),
    (50, "Digital tasks beyond current skill level? / कौशल स्तर से परे?", "radio", ["Yes", "No"]),
    (51, "Training differ by job level? / पद स्तर अनुसार अलग?", "radio", ["Yes", "Somewhat", "No"]),
    (52, "Comfortable asking for help with digital tools / सहायता माँगने में सहज", "likert", ["1","2","3","4","5"]),
    (53, "Org provides adequate support for digital tools / पर्याप्त सहायता", "likert", ["1","2","3","4","5"]),
    (54, "Dept committed to digital transformation / डिजिटल परिवर्तन प्रतिबद्ध", "likert", ["1","2","3","4","5"]),
    (55, "Rank priorities (1=highest, 5=lowest) / प्राथमिकताएँ क्रम दें", "text", None),
    (56, "One change to most improve digital tool use? / सबसे अधिक सुधार?", "text", None),
    (57, "Digital tools improved citizen service delivery? / सेवा वितरण में सुधार?", "radio", ["Yes significantly", "Somewhat", "No change", "Worsened", "Can't say"]),

    # Section 8: Revenue Department Tools
    (58, "[Revenue] Digital tools awareness / राजस्व उपकरण अवगत", "check", ["Bhulekh/WebGIS", "RCMS", "SAARA", "SAMPADA", "e-Court", "None"]),
    (59, "[Revenue] Bhulekh/WebGIS improved land record verification? / भूमि अभिलेख सुधार?", "likert", ["1","2","3","4","5"]),
    (60, "[Revenue] How difficult is RCMS? / RCMS कितना कठिन?", "likert", ["1","2","3","4","5"]),
    (61, "[Revenue] % land records needing paper files / कागज़ी फाइलों की ज़रूरत?", "radio", ["0-20", "21-40", "41-60", "61-80", "81-100"]),
    (62, "[Revenue] Citizens visit expecting digital services? / डिजिटल सेवाओं की अपेक्षा?", "likert", ["1","2","3","4","5"]),
    (63, "[Revenue] Mutation case: which steps digital vs paper? / दाखिल-खारिज डिजिटल vs कागज़?", "text", None),
    (64, "[Revenue] SAMPADA 2.0 reduced registration time? / पंजीकरण समय कम?", "likert", ["1","2","3","4","5"]),

    # Section 9: Rural Development Tools
    (65, "[Rural Dev] Tools awareness / ग्रामीण विकास उपकरण अवगत", "check", ["NREGASoft/NMMS", "e-Gram Swaraj", "PMAY-G", "SBM-G", "Panchayat Darpan", "PFMS", "None"]),
    (66, "[Rural Dev] Difficulty managing multiple portals / कई पोर्टल कठिन?", "likert", ["1","2","3","4","5"]),
    (67, "[Rural Dev] Internet connectivity at block/panchayat / इंटरनेट कनेक्टिविटी", "likert", ["1","2","3","4","5"]),
    (68, "[Rural Dev] NMMS improved MGNREGA attendance accuracy? / उपस्थिति सटीकता?", "likert", ["1","2","3","4","5"]),
    (69, "[Rural Dev] Working day spent on portal data entry / डेटा दर्ज समय", "radio", ["<1hr", "1-2hr", "2-4hr", "4+hr", "Almost all day"]),
    (70, "[Rural Dev] Muster roll/wage list — which steps digital? / कौन-से चरण डिजिटल?", "text", None),
    (71, "[Rural Dev] When e-Gram/NREGASoft down, what do you do? / बंद हो तो क्या करें?", "radio", ["Wait/retry", "Paper", "Upload later", "Block office", "Ask someone", "Skip"]),

    # Section 10: Forest Department Tools
    (72, "[Forest] Digital tools awareness / वन उपकरण अवगत", "check", ["e-Green Watch", "AI Alert", "GIS", "Forest Offence MIS", "Nursery MIS", "None"]),
    (73, "[Forest] AI alert improved illegal activity detection? / अवैध गतिविधि पहचान?", "likert", ["1","2","3","4","5"]),
    (74, "[Forest] How difficult are GIS tools? / GIS कितने कठिन?", "likert", ["1","2","3","4","5"]),
    (75, "[Forest] GPS-enabled device for field verification? / GPS उपकरण?", "radio", ["Dept-issued", "Personal device", "Not available"]),
    (76, "[Forest] When AI alert received, what do you do? / AI अलर्ट पर क्या करें?", "text", None),

    # Section 11: Health Department Tools
    (77, "[Health] Digital tools awareness / स्वास्थ्य उपकरण अवगत", "check", ["ANMOL", "HMIS", "Nikshay", "eVIN", "IHIP", "ABHA", "MPCDSR", "None"]),
    (78, "[Health] ANMOL/ABHA improved patient tracking? / मरीज़ ट्रैकिंग सुधार?", "likert", ["1","2","3","4","5"]),
    (79, "[Health] IHIP reporting adds to workload? / कार्यभार में जोड़ती है?", "likert", ["1","2","3","4","5"]),
    (80, "[Health] ANC registration — which steps on ANMOL? / कौन-से चरण ANMOL पर?", "text", None),
    (81, "[Health] Disease outbreak — how do you report? / रोग प्रकोप रिपोर्ट कैसे?", "text", None),
]

# ── 3. Build the Excel survey workbook ──────────────────────────────────
wb = Workbook()

# --- Sheet 1: Directory with Serial Numbers ---
ws_dir = wb.active
ws_dir.title = "Directory (निर्देशिका)"

# Write header
dir_headers = df.columns.tolist()
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

for col_idx, h in enumerate(dir_headers, 1):
    cell = ws_dir.cell(row=1, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

for row_idx, row_data in df.iterrows():
    for col_idx, val in enumerate(row_data, 1):
        cell = ws_dir.cell(row=row_idx + 2, column=col_idx, value=val)
        cell.border = thin_border

# Auto-width
for col_idx in range(1, len(dir_headers) + 1):
    ws_dir.column_dimensions[get_column_letter(col_idx)].width = 18

# --- Sheet 2: Survey Response Sheet ---
ws_survey = wb.create_sheet("Survey Responses (सर्वे उत्तर)")

# Row 1: Section headers (merged across question ranges)
# Row 2: Question numbers
# Row 3: Question text (bilingual)
# Row 4: Response type / options hint
# Row 5 onward: one row per serial number

# Build section ranges
sections = [
    ("Section 1: Personal Info / व्यक्तिगत जानकारी", 1, 9),
    ("Section 2: Digital Tools Opinion / डिजिटल उपकरणों पर राय", 10, 19),
    ("Section 3: IT Infrastructure / आईटी इंफ्रास्ट्रक्चर", 20, 25),
    ("Section 4: General Usage / सामान्य उपयोग", 26, 35),
    ("Section 5: Other Tools / अन्य उपकरण", 36, 41),
    ("Section 6: Barriers / बाधाएँ", 42, 43),
    ("Section 7: Training / प्रशिक्षण", 44, 57),
    ("Section 8: Revenue / राजस्व", 58, 64),
    ("Section 9: Rural Dev / ग्रामीण विकास", 65, 71),
    ("Section 10: Forest / वन", 72, 76),
    ("Section 11: Health / स्वास्थ्य", 77, 81),
]

# Colors for sections
section_colors = [
    "1B5E20", "2E7D32", "388E3C", "43A047", "4CAF50",
    "66BB6A", "81C784", "A5D6A7", "C8E6C9", "E8F5E9", "B9F6CA"
]

# Column A = Serial No., Column B = Name, Column C = Designation, Column D = Category, Column E = Node
# Columns F onward = questions
fixed_cols = 5  # Serial, Name, Designation, Category, Node

# Row 1: Fixed column headers
fixed_headers = ["Serial No. (क्रमांक)", "Name (नाम)", "Designation (पद)", "Category (श्रेणी)", "Node (नोड)"]
for col_idx, h in enumerate(fixed_headers, 1):
    cell = ws_survey.cell(row=1, column=col_idx, value=h)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")
    cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')
    cell.border = thin_border
    ws_survey.merge_cells(start_row=1, start_column=col_idx, end_row=3, end_column=col_idx)

# Row 1-3: Section headers + Question headers
q_col_start = fixed_cols + 1
for sec_name, sec_start, sec_end in sections:
    sec_col_start = q_col_start + sec_start - 1
    sec_col_end = q_col_start + sec_end - 1
    color_idx = sections.index((sec_name, sec_start, sec_end))
    fill = PatternFill(start_color=section_colors[color_idx], end_color=section_colors[color_idx], fill_type="solid")

    # Merge section header across Row 1
    ws_survey.merge_cells(start_row=1, start_column=sec_col_start, end_row=1, end_column=sec_col_end)
    cell = ws_survey.cell(row=1, column=sec_col_start, value=sec_name)
    cell.font = Font(bold=True, color="FFFFFF", size=10)
    cell.fill = fill
    cell.alignment = Alignment(horizontal='center', wrap_text=True)
    cell.border = thin_border

# Row 2: Question numbers, Row 3: Question text + options hint
for q_num, q_text, q_type, q_options in questions:
    col = q_col_start + q_num - 1

    # Row 2: Q number
    cell2 = ws_survey.cell(row=2, column=col, value=f"Q{q_num}")
    cell2.font = Font(bold=True, size=9)
    cell2.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cell2.alignment = Alignment(horizontal='center')
    cell2.border = thin_border

    # Row 3: Question text + options
    options_hint = ""
    if q_options:
        if q_type == "likert":
            options_hint = " [1-5 scale]"
        elif q_type == "check":
            options_hint = f" [{', '.join(q_options)}]"
        else:
            options_hint = f" [{' | '.join(q_options)}]"

    cell3 = ws_survey.cell(row=3, column=col, value=f"{q_text}{options_hint}")
    cell3.font = Font(size=8)
    cell3.fill = PatternFill(start_color="F1F8E9", end_color="F1F8E9", fill_type="solid")
    cell3.alignment = Alignment(wrap_text=True, vertical='top')
    cell3.border = thin_border

# Row 4 onward: One row per person, LINKED to Directory sheet via formulas
# Directory sheet layout (row 1 = header, row 2+ = data):
#   Col A = Serial No., B = Category, C = Node, D = S.No, E = Name,
#   F = Designation, G = Office Phone, H = Mobile, I = Email, J = Section,
#   K = Additional Charge, L = Fax
# We use INDEX(MATCH()) so the Survey sheet auto-updates when Directory changes.

dir_sheet_name = "'Directory (निर्देशिका)'"
# Directory column letters for lookup (serial is col A = col 1)
# A=Serial, B=Category, C=Node, D=S.No, E=Name, F=Designation, G=OfficePhone, H=Mobile, I=Email
dir_col_map = {
    'name': 'E',        # Name (नाम)
    'designation': 'F', # Designation (पद)
    'category': 'B',    # Category (श्रेणी)
    'node': 'C',        # Node (नोड)
    'mobile': 'H',      # Mobile (मोबाइल)
    'email': 'I',       # Email (ईमेल)
    'office_phone': 'G',# Office Phone
}
total_dir_rows = len(df) + 1  # +1 for header

data_start_row = 4
for idx, row_data in df.iterrows():
    row_num = data_start_row + idx
    serial = int(row_data['Serial No. (क्रमांक)'])

    # Col A: Serial No. (static — this is the lookup key)
    cell = ws_survey.cell(row=row_num, column=1, value=serial)
    cell.border = thin_border

    # Col B: Name — LINKED via INDEX/MATCH
    formula_name = f'=INDEX({dir_sheet_name}!{dir_col_map["name"]}$2:{dir_col_map["name"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=2, value=formula_name)
    cell.border = thin_border

    # Col C: Designation — LINKED
    formula_desig = f'=INDEX({dir_sheet_name}!{dir_col_map["designation"]}$2:{dir_col_map["designation"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=3, value=formula_desig)
    cell.border = thin_border

    # Col D: Category — LINKED
    formula_cat = f'=INDEX({dir_sheet_name}!{dir_col_map["category"]}$2:{dir_col_map["category"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=4, value=formula_cat)
    cell.border = thin_border

    # Col E: Node — LINKED
    formula_node = f'=INDEX({dir_sheet_name}!{dir_col_map["node"]}$2:{dir_col_map["node"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=5, value=formula_node)
    cell.border = thin_border

    # Q1 (Name) — LINKED
    formula_q1 = f'=INDEX({dir_sheet_name}!{dir_col_map["name"]}$2:{dir_col_map["name"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=q_col_start + 0, value=formula_q1)
    cell.border = thin_border

    # Q2 (Designation) — LINKED
    formula_q2 = f'=INDEX({dir_sheet_name}!{dir_col_map["designation"]}$2:{dir_col_map["designation"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=q_col_start + 1, value=formula_q2)
    cell.border = thin_border

    # Q3 (Mobile) — LINKED
    formula_q3 = f'=INDEX({dir_sheet_name}!{dir_col_map["mobile"]}$2:{dir_col_map["mobile"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=q_col_start + 2, value=formula_q3)
    cell.border = thin_border

    # Q4 (Email) — LINKED
    formula_q4 = f'=INDEX({dir_sheet_name}!{dir_col_map["email"]}$2:{dir_col_map["email"]}${total_dir_rows},MATCH($A{row_num},{dir_sheet_name}!$A$2:$A${total_dir_rows},0))'
    cell = ws_survey.cell(row=row_num, column=q_col_start + 3, value=formula_q4)
    cell.border = thin_border

    # Add borders for all remaining question cells
    for q_num, _, _, _ in questions:
        col = q_col_start + q_num - 1
        cell = ws_survey.cell(row=row_num, column=col)
        cell.border = thin_border

# Add data validation (dropdowns) for each question column
for q_num, q_text, q_type, q_options in questions:
    if q_options and q_type in ("radio", "likert"):
        col = q_col_start + q_num - 1
        col_letter = get_column_letter(col)
        formula = '"' + ','.join(q_options) + '"'
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        dv.error = "Please select from the list"
        dv.errorTitle = f"Q{q_num}"
        dv.prompt = f"Select response for Q{q_num}"
        dv.promptTitle = f"Q{q_num}"
        ws_survey.add_data_validation(dv)
        dv.add(f"{col_letter}{data_start_row}:{col_letter}{data_start_row + len(df) - 1}")

# Set column widths
for col_idx in range(1, fixed_cols + 1):
    ws_survey.column_dimensions[get_column_letter(col_idx)].width = 20
for q_num, _, _, _ in questions:
    col = q_col_start + q_num - 1
    ws_survey.column_dimensions[get_column_letter(col)].width = 25

# Freeze panes: freeze first 5 columns and first 3 rows
ws_survey.freeze_panes = "F4"

# Row heights
ws_survey.row_dimensions[1].height = 30
ws_survey.row_dimensions[2].height = 20
ws_survey.row_dimensions[3].height = 80

# Save
output_path = "mp_forest_survey.xlsx"
wb.save(output_path)
print(f"✅ Survey workbook saved to {output_path}")
print(f"   Sheet 1: Directory with {len(df)} serial-numbered entries")
print(f"   Sheet 2: Survey with {len(questions)} questions × {len(df)} respondents")
print(f"   Dropdowns added for {sum(1 for _,_,t,o in questions if o and t in ('radio','likert'))} questions")
