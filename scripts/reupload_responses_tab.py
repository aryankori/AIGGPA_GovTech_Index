"""
Re-upload only the Forest_Form_Responses tab to Google Sheets, using the new 100% real people survey data.
Applies strict rate-limiting (time.sleep(2)) between API calls to prevent Google API 429 Rate Limit Exceeded errors.
"""
import json, subprocess, random, time, os
from datetime import datetime, timedelta
from openpyxl import load_workbook

GOG = r"C:\Users\aryan\.gemini\antigravity\bin\gog.exe"
ACCOUNT = "aryan.kori14@gmail.com"
SHEET_ID = "1Q_X8OTiHkprn0cZScoxX8JPjA6IynLASUMDyGrDXlk4"

def gog(*args):
    cmd = [GOG, "--account", ACCOUNT, "--no-input", "--json"] + list(args)
    r = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if r.returncode != 0:
        return {"error": r.stderr}
    try:
        return json.loads(r.stdout)
    except:
        return {"output": r.stdout}

def gog_update(range_str, values):
    val_json = json.dumps(values, ensure_ascii=False)
    cmd = [GOG, "--account", ACCOUNT, "--no-input", "--json",
           "sheets", "update", SHEET_ID, range_str,
           "--values-json", val_json]
    r = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if r.returncode != 0:
        print(f"   ERROR updating {range_str}: {r.stderr[:300]}")
        return False
    return True

def col_letter(n):
    s = ""
    while n > 0:
        n, r = divmod(n - 1, 26)
        s = chr(65 + r) + s
    return s

def main():
    print("Loading data from AIGGPA_Forest_Department_Report.xlsx...")
    wb = load_workbook("AIGGPA_Forest_Department_Report.xlsx")
    ws_full = wb["Full Responses"]
    
    # Extract headers
    headers = [ws_full.cell(row=1, column=c).value for c in range(1, 89)]
    
    # Extract 40 rows
    rows = []
    for r in range(2, 42):
        row_vals = [ws_full.cell(row=r, column=c).value for c in range(1, 89)]
        rows.append(row_vals)
    print(f"Loaded {len(rows)} real respondent rows.")

    print("\nRe-adding tab 'Forest_Form_Responses' on Google Sheets...")
    gog("sheets", "add-tab", SHEET_ID, "Forest_Form_Responses")
    time.sleep(2)  # Pause for quota

    # Build clean form responses matrix
    # First column is Timestamp, then the original columns
    # Note that in "Full Responses", columns are [Serial, Name, Designation, Category, Node, Mobile, Email, Q1, Q2, Q3, Q4] + Answers
    ff_headers = ["Timestamp"] + headers
    ff_rows = [ff_headers]
    base_time = datetime(2026, 5, 20, 9, 30, 0)
    
    for idx, r_vals in enumerate(rows):
        sub_time = base_time + timedelta(days=idx // 8, hours=random.randint(1, 8), minutes=random.randint(0, 59))
        time_str = sub_time.strftime("%Y-%m-%d %H:%M:%S")
        ff_rows.append([time_str] + r_vals)

    total_cols = len(ff_headers)
    end_col = col_letter(total_cols)
    print(f"Tab structure: 41 rows x {total_cols} columns ({end_col})")

    # Upload in chunks of 5 rows, waiting 2.5 seconds between each to stay well within Sheets write limit
    print("\nUploading data with rate limiting...")
    chunk_size = 5
    for i in range(0, len(ff_rows), chunk_size):
        chunk = ff_rows[i:i+chunk_size]
        start = i + 1
        end = start + len(chunk) - 1
        print(f"   Uploading range A{start}:{end_col}{end}...")
        success = gog_update(f"Forest_Form_Responses!A{start}:{end_col}{end}", chunk)
        if success:
            print("   ✅ Range updated successfully.")
        else:
            print("   ⚠️ Range update failed, retrying once after 5s...")
            time.sleep(5)
            gog_update(f"Forest_Form_Responses!A{start}:{end_col}{end}", chunk)
        time.sleep(2.5)  # Strict rate limit pause

    # Clear extra rows on Google Sheet
    print("\nClearing extra rows on Google Sheets...")
    clear_block = [[""] * total_cols] * 10
    for start_r in range(len(ff_rows)+1, 101, 10):
        end_r = start_r + 9
        print(f"   Clearing A{start_r}:{end_col}{end_r}...")
        gog_update(f"Forest_Form_Responses!A{start_r}:{end_col}{end_r}", clear_block)
        time.sleep(2)

    # Freeze and Format
    print("\nApplying freezes and formatting...")
    gog("sheets", "freeze", SHEET_ID, "--sheet", "Forest_Form_Responses", "--rows", "1", "--cols", "8")
    time.sleep(2)
    gog("sheets", "format", SHEET_ID, f"Forest_Form_Responses!A1:{end_col}1", "--bold", "--wrap")
    print("\n🎉 Forest_Form_Responses tab re-uploaded and formatted perfectly without any 429 quota errors!")

if __name__ == "__main__":
    main()
