import openpyxl
from openpyxl.styles import Alignment, PatternFill
import copy

def add_personal_tools_objective():
    path = r"C:\Users\aryan\Downloads\Framework_for_Interns_updated.xlsx"
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    
    # Find the last row with data
    last_row = ws.max_row
    new_row = last_row + 1
    
    # Use a light purple fill for the new objective
    fill = PatternFill(start_color='FFE1D5E7', end_color='FFE1D5E7', fill_type='solid')
    
    data = {
        1: "5. Personal & Non-Government Tools",
        2: "What non-government tools (WhatsApp, ChatGPT, Google Workspace, YouTube, personal email, etc.) do employees use to support their official work, and do these tools fill gaps left by mandated systems?",
        3: "1. Percentage using non-govt tools for work\n2. Most common tools (WhatsApp, Google Drive, ChatGPT, YouTube)\n3. Primary use cases (drafting, translating, coordinating, learning)\n4. Frequency of personal tool usage\n5. Perceived gap-filling score (Likert)\n6. Data privacy concerns",
        4: "Mixed methods.\nQuantitative: usage frequency, tool counts, Likert gap score.\nQualitative: open-ended privacy concerns.",
        5: "Descriptive statistics (frequencies, percentages).\nCross-tabbing personal tool use by department and age.\nThematic analysis for privacy concern responses.",
        6: "Stacked bar charts for tool adoption by department.\nPie charts for use-case distribution.\nBar charts for gap-filling scores.",
        7: "Head Office and District Office.\nAcross all 4 departments.",
        8: "Q68/Q58. Do you use non-government apps for work?\nQ69/Q59. Which tools? (WhatsApp, Google, ChatGPT, YouTube, etc.)\nQ70/Q60. What for? (Drafting, translating, coordinating, learning, backup)\nQ71/Q61. How often?\nQ72/Q62. Do these fill a gap govt systems don't cover? (1-5)\nQ73/Q63. Privacy or rule concerns?",
        9: "Comparing reported tool use against official IT policy.\nCross-checking with interview data on workaround behaviours.",
        10: "(To be filled after fieldwork)",
        11: "Staff may underreport unofficial tool use if they think it's against rules. Frame questions as neutral - no judgement."
    }
    
    for col, value in data.items():
        cell = ws.cell(row=new_row, column=col)
        cell.value = value
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    wb.save(path)
    print(f"Added 'Personal & Non-Government Tools' objective at row {new_row}.")

if __name__ == "__main__":
    add_personal_tools_objective()
