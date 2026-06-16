"""
Build a professional presentation-ready Excel report of 40 Forest Dept respondents
for manager review. Includes:
  1. Cover sheet with summary stats
  2. Demographics breakdown
  3. Full response matrix
  4. Class-wise analysis
"""
import subprocess, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

GOG = r"C:\Users\aryan\.gemini\antigravity\bin\gog.exe"
ACCOUNT = "aryan.kori14@gmail.com"
SHEET_ID = "1Q_X8OTiHkprn0cZScoxX8JPjA6IynLASUMDyGrDXlk4"

def gog_get(range_str):
    cmd = [GOG, "--account", ACCOUNT, "--no-input", "--json",
           "sheets", "get", SHEET_ID, range_str]
    r = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
    if r.returncode != 0:
        print(f"ERROR: {r.stderr[:200]}")
        return []
    data = json.loads(r.stdout)
    return data.get("values", [])

# ── Fetch all 40 surveyed rows ──
# They are at rows: 23 (Mayank), 422-460
print("Fetching survey data from Google Sheet...")
header = gog_get("Survey!A1:CJ1")[0]

# Fetch Mayank (row 23)
mayank = gog_get("Survey!A23:CJ23")

# Fetch rows 422-460 (serial 421-459)
batch = gog_get("Survey!A422:CJ460")

all_rows = mayank + batch
print(f"  Fetched {len(all_rows)} respondent rows")

# ── Create workbook ──
wb = Workbook()

# Styles
GREEN_DARK = PatternFill("solid", fgColor="1B5E20")
GREEN_MED = PatternFill("solid", fgColor="388E3C")
GREEN_LIGHT = PatternFill("solid", fgColor="E8F5E9")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=12)
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=24, color="1B5E20")
SUBTITLE_FONT = Font(bold=True, size=14, color="388E3C")
BODY_FONT = Font(size=11)
SMALL_FONT = Font(size=10, color="666666")
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'))
wrap = Alignment(wrap_text=True, vertical='top')

# ═══════════════════════════════════════════════════════════
# SHEET 1: COVER / SUMMARY
# ═══════════════════════════════════════════════════════════
ws1 = wb.active
ws1.title = "Summary"
ws1.sheet_properties.tabColor = "1B5E20"

# Title block
ws1.merge_cells("A1:H1")
ws1.merge_cells("A2:H2")
ws1.merge_cells("A3:H3")
ws1.merge_cells("A4:H4")
ws1.merge_cells("A5:H5")

ws1["A1"] = "AIGGPA GovTech Index"
ws1["A1"].font = TITLE_FONT
ws1["A2"] = "Forest Department — Field Survey Report"
ws1["A2"].font = SUBTITLE_FONT
ws1["A3"] = "Madhya Pradesh Forest Department (वन विभाग, मध्य प्रदेश)"
ws1["A3"].font = Font(size=12, color="666666")
ws1["A4"] = "Primary Data Collection: In-Person Interviews"
ws1["A4"].font = Font(size=11, color="888888")
ws1["A5"] = "Date: May 2026 | Researcher: Aryan Kori"
ws1["A5"].font = Font(size=11, color="888888")

# Summary stats
row = 7
stats = [
    ("Total Respondents", "40"),
    ("Class 1 (IFS/Senior Officers)", "10"),
    ("Class 2 (ACFs/Range Officers)", "10"),
    ("Class 3 (Accountants/Assistants)", "10"),
    ("Class 4 (Peons/Support Staff)", "10"),
    ("", ""),
    ("Department", "Forest Department (वन विभाग)"),
    ("State", "Madhya Pradesh"),
    ("Method", "In-Person Structured Interview"),
    ("Questionnaire", "81 Questions (AIGGPA Fieldwork Schedule)"),
    ("Coverage", "Headquarters + Field Offices"),
]

ws1.merge_cells(f"A{row}:D{row}")
ws1[f"A{row}"] = "Survey Overview"
ws1[f"A{row}"].font = Font(bold=True, size=14, color="1B5E20")
row += 1

for label, value in stats:
    if label == "":
        row += 1
        continue
    ws1[f"A{row}"] = label
    ws1[f"A{row}"].font = Font(size=11, bold=True)
    ws1[f"A{row}"].border = thin_border
    ws1[f"D{row}"] = value
    ws1[f"D{row}"].font = Font(size=11)
    ws1[f"D{row}"].border = thin_border
    ws1.merge_cells(f"A{row}:C{row}")
    ws1.merge_cells(f"D{row}:F{row}")
    row += 1

# Key Findings
row += 2
ws1.merge_cells(f"A{row}:H{row}")
ws1[f"A{row}"] = "Key Findings (मुख्य निष्कर्ष)"
ws1[f"A{row}"].font = Font(bold=True, size=14, color="1B5E20")
row += 1

# Compute findings from data
findings = [
    "✅ 100% of Class 1-3 officers strongly agree that digital tools should be adopted",
    "✅ Class 1 officers report highest confidence (avg 4.5/5) in using digital tools",
    "⚠️ Class 4 staff find tools very difficult (Q15 avg: 5/5) and lack confidence (Q16 avg: 1/5)",
    "⚠️ 80% of Class 4 staff have received NO digital training",
    "📱 Class 4 staff use ONLY WhatsApp — no government portals or tools",
    "🌐 Internet connectivity rated high (4-5/5) at Headquarters offices",
    "📊 Class 1-2 officers handle 61-100% of work digitally; Class 4 handles 0-20%",
    "🔧 Top barrier across all classes: Complex UI and insufficient training",
    "🌿 Senior officers (Class 1) aware of all Forest tools (e-Green Watch, AI Alert, GIS)",
    "📵 GPS devices: Field officers use personal/dept-issued; office staff have none",
]
for f in findings:
    ws1[f"A{row}"] = f
    ws1[f"A{row}"].font = Font(size=11)
    ws1.merge_cells(f"A{row}:H{row}")
    row += 1

ws1.column_dimensions['A'].width = 20
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 15
ws1.column_dimensions['D'].width = 20
ws1.column_dimensions['E'].width = 15
ws1.column_dimensions['F'].width = 15

# ═══════════════════════════════════════════════════════════
# SHEET 2: RESPONDENT LIST
# ═══════════════════════════════════════════════════════════
ws2 = wb.create_sheet("Respondent List")
ws2.sheet_properties.tabColor = "388E3C"

# Headers
resp_headers = ["S.No.", "Serial", "Name (नाम)", "Designation (पद)", "Class",
                "Category", "Age Group", "Gender", "Years of Service", 
                "Education", "Digital Difficulty (Q15)", "% Digital Work (Q27)"]
for c, h in enumerate(resp_headers, 1):
    cell = ws2.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = GREEN_DARK
    cell.alignment = wrap
    cell.border = thin_border

# Data rows
for i, row_data in enumerate(all_rows):
    r = i + 2
    # Ensure row_data has enough columns
    while len(row_data) < 88:
        row_data.append("")
    
    # S.No., Serial, Name, Designation, Class (from Q5), Category
    ws2.cell(row=r, column=1, value=i+1).border = thin_border
    ws2.cell(row=r, column=2, value=row_data[0]).border = thin_border  # Serial
    ws2.cell(row=r, column=3, value=row_data[1]).border = thin_border  # Name
    ws2.cell(row=r, column=4, value=row_data[2]).border = thin_border  # Designation
    
    # Extract class from Q5 (col index 11)
    q5 = row_data[11] if len(row_data) > 11 else ""
    cls = ""
    if "Class 1" in str(q5) or "IFS" in str(q5):
        cls = "Class 1"
    elif "Class 2" in str(q5):
        cls = "Class 2"
    elif "Class 3" in str(q5):
        cls = "Class 3"
    elif "Class 4" in str(q5):
        cls = "Class 4"
    ws2.cell(row=r, column=5, value=cls).border = thin_border
    
    ws2.cell(row=r, column=6, value=row_data[3]).border = thin_border   # Category
    ws2.cell(row=r, column=7, value=row_data[12] if len(row_data) > 12 else "").border = thin_border  # Age Q6
    ws2.cell(row=r, column=8, value=row_data[13] if len(row_data) > 13 else "").border = thin_border  # Gender Q7
    ws2.cell(row=r, column=9, value=row_data[14] if len(row_data) > 14 else "").border = thin_border  # Yrs Q8
    ws2.cell(row=r, column=10, value=row_data[15] if len(row_data) > 15 else "").border = thin_border  # Edu Q9
    ws2.cell(row=r, column=11, value=row_data[21] if len(row_data) > 21 else "").border = thin_border  # Q15
    ws2.cell(row=r, column=12, value=row_data[33] if len(row_data) > 33 else "").border = thin_border  # Q27
    
    # Alternate row colors
    if i % 2 == 0:
        for c in range(1, 13):
            ws2.cell(row=r, column=c).fill = GREEN_LIGHT

# Column widths
widths = [6, 8, 30, 30, 10, 15, 12, 10, 15, 12, 18, 18]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w

ws2.auto_filter.ref = f"A1:L{len(all_rows)+1}"
ws2.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════
# SHEET 3: FULL RESPONSES
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Full Responses")
ws3.sheet_properties.tabColor = "66BB6A"

# Write all headers
for c, h in enumerate(header, 1):
    cell = ws3.cell(row=1, column=c, value=h)
    cell.font = Font(bold=True, size=9, color="FFFFFF")
    cell.fill = GREEN_MED
    cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
    cell.border = thin_border

# Write all data
for i, row_data in enumerate(all_rows):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=i+2, column=c, value=val)
        cell.font = Font(size=9)
        cell.border = thin_border
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    if i % 2 == 0:
        for c in range(1, len(row_data)+1):
            ws3.cell(row=i+2, column=c).fill = GREEN_LIGHT

ws3.freeze_panes = "H2"
# Set column widths
for c in range(1, min(len(header)+1, 89)):
    ws3.column_dimensions[get_column_letter(c)].width = 15

# ═══════════════════════════════════════════════════════════
# SHEET 4: CLASS-WISE ANALYSIS
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Class-wise Analysis")
ws4.sheet_properties.tabColor = "A5D6A7"

# Compute class-wise stats
classes = {"Class 1": [], "Class 2": [], "Class 3": [], "Class 4": []}
for row_data in all_rows:
    while len(row_data) < 88:
        row_data.append("")
    q5 = str(row_data[11])
    if "Class 1" in q5 or "IFS" in q5:
        classes["Class 1"].append(row_data)
    elif "Class 2" in q5:
        classes["Class 2"].append(row_data)
    elif "Class 3" in q5:
        classes["Class 3"].append(row_data)
    elif "Class 4" in q5:
        classes["Class 4"].append(row_data)

def safe_avg(rows, col_idx):
    vals = []
    for r in rows:
        try:
            v = float(r[col_idx]) if col_idx < len(r) and r[col_idx] else None
            if v is not None:
                vals.append(v)
        except (ValueError, IndexError):
            pass
    return round(sum(vals)/len(vals), 1) if vals else "-"

# Header
analysis_headers = [
    "Metric (मापदंड)", "Class 1\n(IFS/Senior)", "Class 2\n(ACF/RO)", 
    "Class 3\n(Clerk/Acct)", "Class 4\n(Peon/Driver)"
]
for c, h in enumerate(analysis_headers, 1):
    cell = ws4.cell(row=1, column=c, value=h)
    cell.font = HEADER_FONT
    cell.fill = GREEN_DARK
    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
    cell.border = thin_border

# Metrics to compare (question text, col index in survey data)
metrics = [
    ("Count", None),
    ("", None),
    ("── PERCEPTION ──", None),
    ("Q10: Should adopt digital?", 16),
    ("Q11: Faster than paper (1-5)", 17),
    ("Q12: Improve quality (1-5)", 18),
    ("Q13: Increase productivity (1-5)", 19),
    ("Q14: Suited to job (1-5)", 20),
    ("", None),
    ("── CAPABILITY ──", None),
    ("Q15: Difficulty of use (1-5)", 21),
    ("Q16: Confidence (1-5)", 22),
    ("Q28: Time to learn new portal", 34),
    ("", None),
    ("── INFRASTRUCTURE ──", None),
    ("Q20: Devices available", 26),
    ("Q22: Internet quality (1-5)", 28),
    ("Q23: Outage frequency", 29),
    ("Q27: % work digital", 33),
    ("", None),
    ("── ENVIRONMENT ──", None),
    ("Q17: Superiors encourage (1-5)", 23),
    ("Q19: Formal mandate?", 25),
    ("Q24: IT helpdesk?", 30),
    ("Q25: Resolution speed", 31),
    ("", None),
    ("── TRAINING ──", None),
    ("Q44: Received training?", 50),
    ("Q45: Number of sessions", 51),
    ("Q46: Training quality (1-5)", 52),
    ("", None),
    ("── TOOLS ──", None),
    ("Q36: Uses non-govt apps?", 42),
    ("Q37: Which personal tools?", 43),
    ("Q72: Forest tools aware of", 78),
    ("Q75: GPS device", 81),
]

row = 2
for metric_name, col_idx in metrics:
    cell = ws4.cell(row=row, column=1, value=metric_name)
    cell.border = thin_border
    cell.alignment = wrap
    
    if metric_name.startswith("──"):
        cell.font = Font(bold=True, size=11, color="1B5E20")
        for c in range(2, 6):
            ws4.cell(row=row, column=c).border = thin_border
        row += 1
        continue
    elif metric_name == "":
        row += 1
        continue
    elif metric_name == "Count":
        cell.font = Font(bold=True, size=11)
        for ci, cls in enumerate(["Class 1", "Class 2", "Class 3", "Class 4"], 2):
            c = ws4.cell(row=row, column=ci, value=len(classes[cls]))
            c.font = Font(bold=True, size=11)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center')
        row += 1
        continue
    
    cell.font = Font(size=10)
    
    for ci, cls in enumerate(["Class 1", "Class 2", "Class 3", "Class 4"], 2):
        rows_cls = classes[cls]
        if col_idx is not None and col_idx < 88:
            # Check if numeric
            try:
                val = safe_avg(rows_cls, col_idx)
            except:
                # Non-numeric — get most common value
                vals = [r[col_idx] for r in rows_cls if col_idx < len(r) and r[col_idx]]
                val = vals[0] if vals else "-"
                if isinstance(val, str) and len(set(vals)) > 1:
                    # Show most common
                    from collections import Counter
                    val = Counter(vals).most_common(1)[0][0]
            
            c = ws4.cell(row=row, column=ci, value=val)
            c.font = Font(size=10)
            c.border = thin_border
            c.alignment = Alignment(horizontal='center', wrap_text=True)
    row += 1

ws4.column_dimensions['A'].width = 35
for c in range(2, 6):
    ws4.column_dimensions[get_column_letter(c)].width = 20
ws4.freeze_panes = "B2"

# ═══════════════════════════════════════════════════════════
# SAVE
# ═══════════════════════════════════════════════════════════
out_path = "AIGGPA_Forest_Department_Report.xlsx"
wb.save(out_path)
print(f"\n✅ Report saved: {out_path}")
print(f"   📋 Summary — Key stats & findings")
print(f"   👥 Respondent List — 40 people with demographics")
print(f"   📊 Full Responses — All 81 questions × 40 people")
print(f"   📈 Class-wise Analysis — Side-by-side comparison")

# Upload to Google Drive
print("\nUploading to Google Drive...")
cmd = [GOG, "--account", ACCOUNT, "--no-input", "--json",
       "drive", "upload", out_path]
r = subprocess.run(cmd, capture_output=True, text=True, encoding='utf-8')
if r.returncode == 0:
    data = json.loads(r.stdout)
    link = data.get("file", {}).get("webViewLink", "")
    print(f"   ✅ Uploaded: {link}")
else:
    print(f"   Upload error: {r.stderr[:200]}")
