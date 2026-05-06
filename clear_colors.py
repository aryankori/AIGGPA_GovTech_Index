import openpyxl
from openpyxl.styles import PatternFill

def clear_colors():
    filepath = r"C:\Users\aryan\Downloads\Framework for Interns.xlsx"
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Remove fill from all data rows (row 3 onwards)
    no_fill = PatternFill(fill_type=None)
    
    for row in ws.iter_rows(min_row=3):
        for cell in row:
            cell.fill = no_fill
            
    wb.save(filepath)
    print("Colors removed from data rows.")

if __name__ == "__main__":
    clear_colors()
