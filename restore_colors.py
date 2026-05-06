import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

file_path = r"c:\Users\aryan\Downloads\Framework for Interns_Filled.xlsx"
wb = openpyxl.load_workbook(file_path)
ws = wb["Sheet1"]

# Define the fills based on original workbook
header_fill = PatternFill(patternType='solid', fgColor=Color(theme=5, tint=0.7999816888943144))
obj1_fill = PatternFill(patternType='solid', fgColor=Color(rgb="FFF4F480"))
obj2_fill = PatternFill(patternType='solid', fgColor=Color(theme=8, tint=0.7999816888943144))
obj3_fill = PatternFill(patternType='solid', fgColor=Color(theme=9, tint=0.5999938962981048))
# Add a 4th color for the 4th objective (theme 7, tint 0.8 is typically a light green)
obj4_fill = PatternFill(patternType='solid', fgColor=Color(theme=7, tint=0.7999816888943144))

fills = [obj1_fill, obj2_fill, obj3_fill, obj4_fill]

# Row mappings from our filled script:
# Header: row 3
# Obj 1: rows 4-7
# Obj 2: rows 8-10
# Obj 3: rows 11-13
# Obj 4: rows 14-16

# Apply header fill
for c in range(1, 12):
    ws.cell(row=3, column=c).fill = header_fill

# Apply data fills
def apply_fill(start_row, end_row, fill):
    for r in range(start_row, end_row + 1):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = fill

apply_fill(4, 7, obj1_fill)
apply_fill(8, 10, obj2_fill)
apply_fill(11, 13, obj3_fill)
apply_fill(14, 16, obj4_fill)

wb.save(file_path)
print("Colors restored successfully.")
