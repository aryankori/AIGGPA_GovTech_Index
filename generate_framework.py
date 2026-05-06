"""
Script to populate the 'Framework for Interns.xlsx' based on
the Concept Proposal (Wellness_Himani_V1.docx).

Maps research objectives → research questions → indicators →
method type → analysis tools → visuals → levels → questionnaire items →
means of verification, following the proposal's structure.
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from copy import copy

SRC = r"c:\Users\aryan\Downloads\Framework for Interns.xlsx"
OUT = r"c:\Users\aryan\Downloads\Framework for Interns_Filled.xlsx"

wb = openpyxl.load_workbook(SRC)
ws = wb["Sheet1"]

# ── 0. Unmerge existing merged cells so we can write freely ──────────────
for mc in list(ws.merged_cells.ranges):
    ws.unmerge_cells(str(mc))

# ── 1. Title Row ─────────────────────────────────────────────────────────
ws["B1"] = "Title of the Research/Study"
ws["C1"] = (
    "Beyond the Desk: Wellness Dimensions and Their Impact on "
    "Efficiency and Performance in Government Workplaces"
)
ws["C1"].font = Font(bold=True, size=12)
ws.merge_cells("C1:K1")

# ── 2. Header Row (row 3) — already present, just ensure bold ───────────
headers = [
    "Research Objectives",           # A
    "Research Questions",            # B
    "Indicators",                    # C
    "Quantitative/ Qualitative",     # D
    "Analysis Tools",                # E
    "Visual Representation",         # F
    "Levels",                        # G
    "Schedule/Questionnaire Questions",  # H
    "Means of Verification",        # I
    "Findings",                      # J
    "Doubts",                        # K
]
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

for col_idx, h in enumerate(headers, start=1):
    cell = ws.cell(row=3, column=col_idx, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = thin_border

# ── 3. Research data (from the proposal) ─────────────────────────────────
# The proposal has 4 research objectives. The framework template had 3
# merged groups (rows 4-7, 8-11, 12-13). We'll expand to 4 objectives,
# each getting multiple rows for sub-indicators.

data = [
    # ── Objective 1 ──────────────────────────────────────────────────────
    {
        "objective": (
            "RO1: Evaluate the current wellness status of government "
            "employees to understand the need for structured wellness programs."
        ),
        "rows": [
            {
                "question": "What is the current state of physical, emotional, social, and environmental wellness among government employees?",
                "indicator": "Self-reported physical health, sleep quality, diet, exercise frequency, BMI",
                "method": "Quantitative",
                "tools": "Descriptive Statistics (Mean, SD, Frequency); Likert-scale analysis",
                "visual": "Bar charts; Radar/Spider charts across wellness dimensions",
                "levels": "Class I, II, III, IV employees across 5 departments",
                "questions": "How often do you engage in physical activity per week? Rate your sleep quality (1–5). Rate your overall stress level (1–10).",
                "verification": "Structured questionnaire responses; Biometric screening data (BMI, BP)",
            },
            {
                "question": "",
                "indicator": "Self-reported stress levels (1–10 scale); Burnout risk assessment scores",
                "method": "Quantitative",
                "tools": "Descriptive Statistics; Burnout Inventory Scoring",
                "visual": "Heatmaps of stress by department; Distribution histograms",
                "levels": "",
                "questions": "How would you rate your ability to cope with daily work pressures? Have you ever used counselling or EAP services?",
                "verification": "Questionnaire responses; EAP utilization records (if available)",
            },
            {
                "question": "",
                "indicator": "Workplace safety & ergonomics; Office infrastructure satisfaction; Access to clean water & sanitation",
                "method": "Quantitative",
                "tools": "Frequency analysis; Satisfaction index",
                "visual": "Stacked bar charts (satisfaction levels)",
                "levels": "",
                "questions": "Rate your satisfaction with office infrastructure — lighting, ventilation, cleanliness (1–5). Is a dedicated break/wellness room available?",
                "verification": "Questionnaire responses; Physical facility audit checklist",
            },
            {
                "question": "",
                "indicator": "Sense of purpose; Engagement in mindfulness/yoga; Community service participation",
                "method": "Quantitative",
                "tools": "Frequency & percentage analysis",
                "visual": "Pie charts; Grouped bar charts",
                "levels": "",
                "questions": "Do you feel a sense of purpose in your daily work? How often do you engage in mindfulness, meditation, or yoga?",
                "verification": "Questionnaire responses; AYUSH programme participation records",
            },
        ],
    },
    # ── Objective 2 ──────────────────────────────────────────────────────
    {
        "objective": (
            "RO2: Understand how absence of wellness affects the day-to-day "
            "performance of employees, and identify where support is most "
            "needed across different roles."
        ),
        "rows": [
            {
                "question": "How does employee wellness (or its absence) impact day-to-day efficiency, absenteeism, and presenteeism?",
                "indicator": "Sick leave & absenteeism rates; Presenteeism self-assessment scores",
                "method": "Quantitative",
                "tools": "Correlation analysis (Pearson/Spearman); Regression analysis",
                "visual": "Scatter plots (wellness score vs. absenteeism); Line graphs",
                "levels": "Class I, II, III, IV employees across 5 departments",
                "questions": "How many sick leaves have you taken in the last 6 months? How often do you come to work despite feeling unwell? Rate how health issues affect your decision-making speed (1–5).",
                "verification": "Questionnaire responses; Leave records from department HR",
            },
            {
                "question": "",
                "indicator": "Job satisfaction survey scores; Work-life balance self-assessment; Manager support feedback",
                "method": "Quantitative",
                "tools": "Likert-scale analysis; ANOVA (across employee classes)",
                "visual": "Box plots (satisfaction by class); Diverging bar charts",
                "levels": "",
                "questions": "Rate your overall job satisfaction (1–5). Do you feel you have a healthy work-life balance? How supported do you feel by your manager/supervisor?",
                "verification": "Questionnaire responses; Retention & turnover data",
            },
            {
                "question": "",
                "indicator": "Financial stress impact on work focus; Awareness of welfare entitlements (NPS, GPF)",
                "method": "Quantitative",
                "tools": "Frequency analysis; Cross-tabulation",
                "visual": "Grouped bar charts; Frequency tables",
                "levels": "",
                "questions": "Does financial stress affect your ability to focus at work? Are you aware of employee welfare entitlements like NPS and GPF?",
                "verification": "Questionnaire responses; NPS/GPF enrollment data",
            },
        ],
    },
    # ── Objective 3 ──────────────────────────────────────────────────────
    {
        "objective": (
            "RO3: Explore how age, gender and responsibilities shape the "
            "wellness needs of government personnel, so that support can be "
            "more focused and useful for different groups."
        ),
        "rows": [
            {
                "question": "How do demographic factors (age, gender, employee class) shape the wellness needs and priorities of government employees?",
                "indicator": "Wellness dimension scores segmented by age, gender, class",
                "method": "Quantitative",
                "tools": "Chi-square test; Independent t-test; ANOVA; Cross-tabulation",
                "visual": "Grouped bar charts (gender × wellness); Box plots (age groups); Heatmaps",
                "levels": "Class I, II, III, IV — segmented by age group and gender",
                "questions": "Demographic section: Age, Gender, Employee Class, Department, Years of service. Which wellness area do you feel needs the most attention?",
                "verification": "Questionnaire demographic data; GAD classification records",
            },
            {
                "question": "What are the lived experiences and perceived wellness challenges of employees across different levels?",
                "indicator": "Qualitative themes: hierarchy-related stress, peer-support, organizational culture",
                "method": "Qualitative",
                "tools": "Thematic analysis; Content analysis of interview transcripts",
                "visual": "Word clouds; Thematic maps; Quote tables",
                "levels": "Select officials from different classes and departments",
                "questions": "Describe the most stressful part of your workday. How does your current state of health impact your decision-making speed? What kind of wellness support would you most appreciate?",
                "verification": "Semi-structured interview recordings & transcripts; Field notes",
            },
            {
                "question": "",
                "indicator": "Interpersonal relationship quality; Sense of belonging; Peer recognition frequency",
                "method": "Mixed",
                "tools": "Likert-scale analysis; Narrative coding",
                "visual": "Radar charts (social wellness by class); Quote excerpts",
                "levels": "",
                "questions": "Rate the quality of your relationships with colleagues (1–5). Do you feel a sense of belonging at your workplace? How often do you participate in team-building events?",
                "verification": "Questionnaire responses; Interview transcripts",
            },
        ],
    },
    # ── Objective 4 ──────────────────────────────────────────────────────
    {
        "objective": (
            "RO4: Develop an evidence-based framework that demonstrates "
            "the correlation between holistic wellness investments and "
            "increased employee efficiency."
        ),
        "rows": [
            {
                "question": "What evidence-based interventions and frameworks can be recommended to improve government employee wellness and performance?",
                "indicator": "Composite Wellness Index (across 8 dimensions); Correlation with performance metrics",
                "method": "Quantitative",
                "tools": "Multiple regression analysis; Composite index construction; Benchmarking against WHO model",
                "visual": "Framework diagram; Correlation matrix heatmap; Before-after comparison charts",
                "levels": "Aggregated across all sampled employees",
                "questions": "Overall wellness composite derived from all dimension responses. Performance self-assessment: Rate your efficiency at work over the past month (1–5).",
                "verification": "Aggregated questionnaire data; Statistical model outputs; Cross-reference with WHO Healthy Workplace Framework & SAMHSA dimensions",
            },
            {
                "question": "How can existing government initiatives (AYUSH, ABHA, NMHP, Mission Karmayogi) be leveraged for structured wellness programs?",
                "indicator": "Awareness & utilization of government wellness initiatives; Perceived effectiveness of existing programs",
                "method": "Mixed",
                "tools": "Frequency analysis; Gap analysis; Thematic analysis of qualitative responses",
                "visual": "Gap analysis charts; Comparison tables; Policy mapping diagrams",
                "levels": "All employee classes",
                "questions": "Are you aware of AYUSH wellness initiatives? Have you used ABHA (digital health records)? How effective do you find existing health support from the government?",
                "verification": "Questionnaire responses; ABHA enrollment data; AYUSH programme reports; Interview transcripts",
            },
            {
                "question": "",
                "indicator": "Participation in professional development & training; Engagement in creative problem-solving",
                "method": "Quantitative",
                "tools": "Frequency analysis; Satisfaction index",
                "visual": "Bar charts; Trend lines",
                "levels": "",
                "questions": "Have you participated in any professional development/training in the past year? Rate your satisfaction with opportunities for skill development (1–5).",
                "verification": "Training records; Questionnaire responses",
            },
        ],
    },
]

# ── 4. Write data rows ──────────────────────────────────────────────────
wrap_align = Alignment(vertical="top", wrap_text=True)
data_font = Font(size=10)
obj_font = Font(size=10, bold=True)

current_row = 4
col_map = {
    "question": 2,      # B
    "indicator": 3,      # C
    "method": 4,         # D
    "tools": 5,          # E
    "visual": 6,         # F
    "levels": 7,         # G
    "questions": 8,      # H
    "verification": 9,   # I
}

# Alternate objective colors
obj_colors = ["D6E4F0", "E2EFDA", "FCE4D6", "EDEDED"]

for obj_idx, obj_block in enumerate(data):
    start_row = current_row
    objective_text = obj_block["objective"]
    rows_data = obj_block["rows"]
    fill_color = PatternFill(
        start_color=obj_colors[obj_idx],
        end_color=obj_colors[obj_idx],
        fill_type="solid",
    )

    for row_data in rows_data:
        for key, col in col_map.items():
            cell = ws.cell(row=current_row, column=col, value=row_data.get(key, ""))
            cell.alignment = wrap_align
            cell.font = data_font
            cell.border = thin_border
            cell.fill = fill_color

        # Findings & Doubts columns — leave blank for now
        for extra_col in [10, 11]:  # J, K
            cell = ws.cell(row=current_row, column=extra_col)
            cell.border = thin_border
            cell.fill = fill_color
            cell.alignment = wrap_align

        current_row += 1

    end_row = current_row - 1

    # Write objective in column A (merged across rows for this objective)
    ws.cell(row=start_row, column=1, value=objective_text).font = obj_font
    ws.cell(row=start_row, column=1).alignment = Alignment(
        vertical="center", wrap_text=True
    )
    ws.cell(row=start_row, column=1).border = thin_border
    ws.cell(row=start_row, column=1).fill = fill_color
    if end_row > start_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=end_row,
            end_column=1,
        )
        # Apply border/fill to all merged cells
        for r in range(start_row + 1, end_row + 1):
            ws.cell(row=r, column=1).border = thin_border
            ws.cell(row=r, column=1).fill = fill_color

# ── 5. Findings & Doubts note ───────────────────────────────────────────
note_row = current_row + 1
ws.cell(row=note_row, column=10, value="(To be filled after data collection)").font = Font(
    italic=True, color="808080", size=9
)
ws.cell(row=note_row, column=11, value="(To be filled during analysis)").font = Font(
    italic=True, color="808080", size=9
)

# ── 6. Keep the side note about Likert scale (rows 13-15 cols M-N) ─────
# These were in the original file; let's reposition them below the data
scale_start = current_row + 2
ws.cell(row=scale_start, column=12, value="Likert Scale Reference:").font = Font(
    bold=True, size=10
)
scale_data = [
    ("1 – No improvement", "No noticeable change"),
    ("2 – Slightly improved", "Minor, barely noticeable improvement"),
    ("3 – Moderately improved", "Noticeable but not dramatic improvement"),
    ("4 – Significantly improved", "Clear, substantial improvement"),
    ("5 – Greatly improved", "Major, transformative improvement"),
]
for i, (point, meaning) in enumerate(scale_data):
    ws.cell(row=scale_start + 1 + i, column=12, value=point).font = Font(size=9)
    ws.cell(row=scale_start + 1 + i, column=13, value=meaning).font = Font(size=9)

# Clear old scale note cells (M13:N15) if they are outside our data range
for r in range(13, 16):
    for c in range(13, 15):
        old = ws.cell(row=r, column=c)
        if r < start_row or r > end_row:  # don't clear if overlapping data
            old.value = None
# Also clear K13 old note
ws.cell(row=13, column=11).value = None

# ── 7. Column widths ────────────────────────────────────────────────────
col_widths = {
    "A": 35,  # Objectives
    "B": 40,  # Research Questions
    "C": 35,  # Indicators
    "D": 18,  # Quant/Qual
    "E": 32,  # Analysis Tools
    "F": 30,  # Visual Representation
    "G": 28,  # Levels
    "H": 45,  # Questionnaire Questions
    "I": 35,  # Means of Verification
    "J": 20,  # Findings
    "K": 20,  # Doubts
    "L": 22,  # Likert label
    "M": 28,  # Scale Point
    "N": 32,  # Meaning
}
for col_letter, width in col_widths.items():
    ws.column_dimensions[col_letter].width = width

# Row heights
for r in range(4, current_row):
    ws.row_dimensions[r].height = 60

# ── 8. Save ──────────────────────────────────────────────────────────────
wb.save(OUT)
print(f"Done! Saved to: {OUT}")
print(f"Total data rows: {current_row - 4}")
