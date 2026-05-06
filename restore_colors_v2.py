import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.colors import Color

file_path = r"c:\Users\aryan\Downloads\Framework for Interns_Filled.xlsx"
out_path = r"c:\Users\aryan\Downloads\Framework for Interns_Filled_Colors.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
except Exception as e:
    print(f"Error loading workbook: {e}")
    exit(1)

ws = wb["Sheet1"]

# Define the fills based on original workbook
header_fill = PatternFill(patternType='solid', fgColor=Color(theme=5, tint=0.7999816888943144))
obj1_fill = PatternFill(patternType='solid', fgColor=Color(rgb="FFF4F480"))
obj2_fill = PatternFill(patternType='solid', fgColor=Color(theme=8, tint=0.7999816888943144))
obj3_fill = PatternFill(patternType='solid', fgColor=Color(theme=9, tint=0.5999938962981048))
obj4_fill = PatternFill(patternType='solid', fgColor=Color(theme=7, tint=0.7999816888943144))

# Apply header fill
for c in range(1, 12):
    ws.cell(row=3, column=c).fill = header_fill

# Apply data fills
def apply_fill(start_row, end_row, fill):
    for r in range(start_row, end_row + 1):
        for c in range(1, 12):
            ws.cell(row=r, column=c).fill = fill

apply_fill(4, 7, obj1_fill)
apply_fill(8, 10, obj2_fill)
apply_fill(11, 13, obj3_fill)
apply_fill(14, 16, obj4_fill)

try:
    wb.save(out_path)
    print(f"Colors restored successfully. Saved to: {out_path}")
except Exception as e:
    print(f"Failed to save file: {e}")
