import openpyxl
from openpyxl.styles import Alignment, Font, Border, Side

def format_framework():
    filepath = r"C:\Users\aryan\Downloads\Framework for Interns.xlsx"
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    # Unmerge all cells first
    for merged_cell in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged_cell))
        
    # Title
    ws["B1"] = "Title of the Research/Study: Assessment of Digital Tool Adoption and Its Impact on Efficiency in MP Government Departments (Revenue, Rural Development, Forest, Health)"
    ws["B1"].font = Font(bold=True, size=12)
    
    # We will write data starting from row 3 (assuming row 2 has headers)
    # Headers are:
    # A: Research Objectives
    # B: Research Questions
    # C: Indicators
    # D: Quantitative/ Qualitative
    # E: Analysis Tools
    # F: Visual Representation
    # G: Levels
    # H: Schedule/Questionnaire Questions
    # I: Means of Verification
    # J: Findings
    # K: Doubts
    
    data = [
        [
            "1. Digital Infrastructure & Awareness",
            "What is the current status of digital infrastructure in the selected departments, and how aware are employees of the tools available to them?",
            "1. Number and type of devices per employee\n2. Internet bandwidth and uptime\n3. Awareness percentage for key portals\n4. Daily usage frequency\n5. Specific tools used (e-Office, RCMS, etc.)",
            "Mixed methods.\nQuantitative: device counts, usage scores.\nQualitative: field observation notes.",
            "Descriptive statistics (means, frequencies).\nCross-tabbing by department and cadre.\nMapping to TAM 'Facilitating Conditions'.",
            "Bar charts for device availability.\nHeatmaps showing tool usage across cadres.\nStacked bars for awareness by age.",
            "Head Office and District Office.\nAcross all 4 departments.",
            "Q1. What devices are at your workstation?\nQ2. Rate your internet connectivity (1-5).\nQ3. Which of these tools do you know about?\nQ4. How often do you use them?\nQ5. (Observation) Physical setup notes.",
            "Direct office observation.\nChecking department IT inventories.\nComparing survey answers with what we actually see.",
            "(To be filled after fieldwork)",
            "How should we count shared devices? Per device or per user?"
        ],
        [
            "2. Capacity Building & Training",
            "What training and technical support do departments provide to help staff use digital tools?",
            "1. Training programs held in the last 2 years\n2. Percentage of staff who attended formal IT training\n3. Training duration and content\n4. Availability of an IT helpdesk\n5. Staff satisfaction with training quality",
            "Mixed methods.\nQuantitative: attendance rates, satisfaction scores.\nQualitative: interview feedback on training quality.",
            "Descriptive stats.\nLikert-scale averages.\nThematic analysis for interview transcripts.\nMapping to UTAUT 'Effort Expectancy'.",
            "Grouped bar charts for training by department.\nPie charts for trained vs. untrained staff.\nRadar charts for training quality metrics.",
            "Head Office and District Office.\nAcross all 4 departments.",
            "Q6. Have you received formal IT training?\nQ7. How many sessions in the last 2 years?\nQ8. Rate the quality and relevance (1-5).\nQ9. Is there an IT helpdesk here?\nQ10. (Interview) What was useful about the training, and what was missing?",
            "Official training circulars.\niGOT/Mission Karmayogi enrollment data.\nCross-checking staff claims against official records.",
            "(To be filled after fieldwork)",
            "Departments might not keep good training records. How do we verify attendance if HR doesn't have the data?"
        ],
        [
            "3. Bottlenecks & Challenges",
            "What are the main roadblocks employees face when using digital tools in their daily work?",
            "1. Common technical issues (internet, crashes)\n2. System downtime frequency\n3. Perceived difficulty of the tools\n4. Attitudinal barriers (resistance to change)\n5. Organizational barriers (no clear mandate)",
            "Mixed methods.\nQuantitative: issue frequency, difficulty ratings.\nQualitative: in-depth interview and FGD narratives.",
            "Thematic analysis (main approach).\nFrequency counts for reported issues.\nRanking analysis to prioritize problems.",
            "Pareto charts for top bottlenecks.\nWord clouds from interview transcripts.\nDiverging bar charts for difficulty ratings.",
            "Head Office and District Office.\nAll 4 departments.\nAll Cadres (I to IV).",
            "Q11. What are your biggest challenges with digital tools?\nQ12. How often do you face technical issues?\nQ13. How difficult are the tools to use (1-5)?\nQ14. Do you get enough support when things break?\nQ15. (FGD) What systemic changes would make things easier?",
            "Comparing survey, interview, and FGD data to find common patterns.\nObserving actual disruptions during visits.\nChecking IT complaint logs if available.",
            "(To be filled after fieldwork)",
            "Staff might underreport issues because they fear backlash from management. We need strict anonymity protocols."
        ],
        [
            "4. Recommendations & Actions",
            "What specific steps can we suggest to improve digital tool usage and public service delivery?",
            "1. Priority improvement areas ranked by staff\n2. Specific software/hardware requests\n3. Training gap assessments\n4. Infrastructure upgrade needs\n5. Proposed policy changes",
            "Primarily qualitative.\nBuilt directly from the results of Objectives 1 through 3.",
            "Gap analysis (current vs. desired state).\nPriority matrix mapping.\nComparing our findings against standard benchmarks.",
            "Impact vs. Feasibility priority matrix.\nSummary tables for recommendations.\nTimeline roadmap for short/medium/long-term actions.",
            "Department-level interventions.\nState-level policy recommendations.",
            "Q16. What specific improvements do you need?\nQ17. If you could pick one area to fix, what would it be?\nQ18. (Interview) In an ideal world, how would digital tools support your job?\nQ19. (FGD) What policy changes are most urgent?",
            "Ensuring all recommendations are backed by hard data from Objectives 1-3.\nChecking feasibility against known budget constraints.",
            "(To be filled after fieldwork)",
            "Since this relies on the other objectives, we can't finalize it until fieldwork is done. Should we draft preliminary ideas based on literature first?"
        ]
    ]
    
    # Write data
    start_row = 3
    for row_idx, row_data in enumerate(data, start=start_row):
        for col_idx, cell_value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = cell_value
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            
    # Adjust column widths for better readability
    for col in ws.columns:
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = 30
        
    wb.save(filepath)
    print("Successfully formatted and populated Framework for Interns.xlsx")

if __name__ == "__main__":
    format_framework()
